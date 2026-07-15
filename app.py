from flask import Flask, render_template, request, jsonify, Response, send_file, render_template_string
from flask_socketio import SocketIO, emit
import sqlite3
import requests
from datetime import datetime, timedelta
import io
import csv
import urllib.parse
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
import os

# ═══════════════════════════════════════════════
# 🔔 PUSH NOTIFICATIONS — VAPID CONFIG
# ═══════════════════════════════════════════════
try:
    from pywebpush import webpush, WebPushException
    from py_vapid import Vapid
    PUSH_ENABLED = True
except ImportError:
    PUSH_ENABLED = False
    print("[push] pywebpush not installed — push disabled")

# VAPID keys — generated once, stored in vapid_private.pem next to app.py
VAPID_PRIVATE_KEY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vapid_private.pem")
# لو في Railway — اكتب الـ PEM من الـ env variable
_vapid_b64 = os.environ.get("VAPID_PRIVATE_KEY_B64")
if _vapid_b64:
    import base64 as _b64mod
    with open(VAPID_PRIVATE_KEY_PATH, "wb") as _vf:
        _vf.write(_b64mod.b64decode(_vapid_b64))
VAPID_PUBLIC_KEY = "BITUMkU5H85CCfFNQlGxDZuJcGWJBR4a5qF8-Wndf9s-s96TRmgawjU4ArNU53XSTH486lSaOdTyJdM9vcYYEPs"
VAPID_CLAIMS = {"sub": "mailto:admin@rwa7el.local"}

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")
DB = "queue.db"
def con():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c
with con() as db:
    db.execute("""CREATE TABLE IF NOT EXISTS drivers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        status TEXT DEFAULT 'waiting',
        created TEXT,
        out_time TEXT,
        done_time TEXT,
        sort_order INTEGER DEFAULT 0,
        break_start TEXT,
        break_paused INTEGER DEFAULT 0,
        break_paused_accum INTEGER DEFAULT 0,
        break_pause_started TEXT,
        misses INTEGER DEFAULT 0,
        returns INTEGER DEFAULT 0,
        lat REAL,
        lng REAL,
        battery INTEGER,
        charging INTEGER DEFAULT 0
    )""")
    for col in ["sort_order INTEGER DEFAULT 0", "break_start TEXT", "break_paused INTEGER DEFAULT 0", 
                 "break_paused_accum INTEGER DEFAULT 0", "break_pause_started TEXT",
                 "misses INTEGER DEFAULT 0", "returns INTEGER DEFAULT 0", "lat REAL", "lng REAL",
                 "battery INTEGER", "charging INTEGER DEFAULT 0"]:
        try:
            db.execute(f"ALTER TABLE drivers ADD COLUMN {col}")
        except:
            pass
    db.execute("UPDATE drivers SET sort_order = id WHERE sort_order = 0")
    db.execute("""CREATE TABLE IF NOT EXISTS settings(
        key TEXT PRIMARY KEY,
        value TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS driver_orders(
        name TEXT PRIMARY KEY,
        orders INTEGER DEFAULT 0,
        last_done TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS miss_history(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        driver_name TEXT,
        order_id TEXT,
        reason TEXT,
        timestamp TEXT,
        date TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS return_history(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        driver_name TEXT,
        order_id TEXT,
        timestamp TEXT,
        date TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS done_history(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        driver_name TEXT,
        timestamp TEXT,
        date TEXT
    )""")
    for tbl, col in [("miss_history", "date TEXT"), ("return_history", "date TEXT"), ("done_history", "date TEXT")]:
        try:
            db.execute(f"ALTER TABLE {tbl} ADD COLUMN {col}")
        except:
            pass
    db.execute("""CREATE TABLE IF NOT EXISTS allowed_accounts(
        email TEXT PRIMARY KEY
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS dispatcher_orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_hash TEXT UNIQUE,
        time TEXT,
        distance TEXT,
        items TEXT,
        weight TEXT,
        pickup TEXT,
        dropoff TEXT,
        picked_up_by TEXT,
        created_at TEXT
    )""")
    try:
        db.execute("ALTER TABLE dispatcher_orders ADD COLUMN picked_up_by TEXT")
    except:
        pass
    db.execute("""CREATE TABLE IF NOT EXISTS geocode_cache(
        address TEXT PRIMARY KEY,
        lat REAL,
        lng REAL,
        cached_at TEXT
    )""")
    # 🗺️ إحداثيات دقيقة اتجابت من جوجل مابس (عن طريق يوزرسكريبت GoogleMapsGeocode.txt)
    # بتتفضل ليها الأولوية على أي geocoding نصي عادي (Nominatim) لأنها أدق بكتير
    db.execute("""CREATE TABLE IF NOT EXISTS resolved_locations(
        address TEXT PRIMARY KEY,
        lat REAL,
        lng REAL,
        resolved_at TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS shift_state(
        id INTEGER PRIMARY KEY CHECK (id = 1),
        started_at TEXT,
        started_at_display TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS fines(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        driver_name TEXT,
        amount REAL,
        reason TEXT,
        timestamp TEXT,
        date TEXT
    )""")
    # ربط الإيميل (اللي بيستخدمه نظام الكيو كـ name) بالاسم الحقيقي لنظام التقييم
    # 🔔 Push notification subscriptions — مفتاحها اسم الطيار
    db.execute("""CREATE TABLE IF NOT EXISTS push_subscriptions(
        name TEXT PRIMARY KEY,
        subscription TEXT,
        updated_at TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS name_mapping(
        email TEXT PRIMARY KEY,
        full_name TEXT
    )""")
    # ملاحظات نظام التقييم لكل مندوب (مفتاحها الإيميل = نفس name في باقي الجداول)
    db.execute("""CREATE TABLE IF NOT EXISTS driver_rating_notes(
        email TEXT PRIMARY KEY,
        notes TEXT,
        updated_at TEXT
    )""")
    db.commit()

def seed_allowed_accounts_from_xlsx():
    import os
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rwa7el_accounts.xlsx")
    if not os.path.exists(xlsx_path):
        return
    db = con()
    try:
        existing = db.execute("SELECT COUNT(*) as c FROM allowed_accounts").fetchone()["c"]
        if existing > 0:
            return
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        emails = set()
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell is None: continue
                val = str(cell).strip().lower()
                if "@" in val and val != "email address":
                    emails.add(val)
                    break
        if emails:
            db.executemany("INSERT OR IGNORE INTO allowed_accounts(email) VALUES(?)", [(e,) for e in emails])
            db.commit()
    except Exception as e:
        pass
    finally:
        db.close()

seed_allowed_accounts_from_xlsx()

def seed_name_mapping_from_xlsx():
    """يقرأ rwa7el_name_mapping.xlsx (لو موجود جنب app.py) ويعمل seed مرة واحدة بس لو الجدول فاضي.
    العمود الأول = الاسم الحقيقي، العمود التاني = الإيميل."""
    import os
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rwa7el_name_mapping.xlsx")
    if not os.path.exists(xlsx_path):
        return
    db = con()
    try:
        existing = db.execute("SELECT COUNT(*) as c FROM name_mapping").fetchone()["c"]
        if existing > 0:
            return
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        pairs = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            name_cell, email_cell = row[0], row[1]
            if not name_cell or not email_cell:
                continue
            name = str(name_cell).strip()
            email = str(email_cell).strip().lower()
            if "@" not in email or name.lower() == "name":
                continue
            pairs.append((email, name))
        if pairs:
            db.executemany("INSERT OR IGNORE INTO name_mapping(email, full_name) VALUES(?,?)", pairs)
            db.commit()
    except Exception:
        pass
    finally:
        db.close()

seed_name_mapping_from_xlsx()

def now():
    return datetime.now().strftime("%H:%M:%S")

def today():
    return datetime.now().strftime("%Y-%m-%d")

def public_url():
    # 1) ngrok tunnel (priority - accessible from any network)
    try:
        data = requests.get("http://127.0.0.1:4040/api/tunnels", timeout=1).json()
        for t in data["tunnels"]:
            if t["public_url"].startswith("https://"):
                return t["public_url"]
    except:
        pass
    # 2) Local network IP (works when admin + drivers on same WiFi)
    try:
        import socket as _socket
        s = _socket.socket(_socket.AF_INET, _socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        if not local_ip.startswith("127."):
            return f"http://{local_ip}:5000"
    except:
        pass
    return "http://127.0.0.1:5000"

# ═══════════════════════════════════════════════════════════
# 🗺️ GEOCODING — يحول عنوان نصي لإحداثيات (lat/lng)
# باستخدام OpenStreetMap Nominatim (مجاني، محدود لطلب واحد/ثانية)
# 🔴 من غير كاش خالص — كل نداء بيعمل geocode حي (live) من الصفر
# (الاستثناء: Plus Code بيتفك محليًا فورًا، مش نداء شبكة أصلاً)
# ═══════════════════════════════════════════════════════════
import math, time as _time
from openlocationcode import openlocationcode as olc
_last_geocode_call = [0.0]  # آخر وقت اتنادى فيه Nominatim فعليًا (مش من الكاش)
GEOCODE_MIN_INTERVAL = 1.05  # ثانية بين كل طلب حقيقي، احترامًا لسياسة استخدام Nominatim

# الإسكندرية تقريبًا بين خطي عرض 31.0 و31.35، وخطي طول 29.7 و30.15
# بنستخدمهم كـ viewbox عشان نقيّد نتائج Nominatim جغرافيًا ومنجيبش نتيجة غلط من بلد/مدينة تانية
ALEXANDRIA_VIEWBOX = "29.70,31.40,30.25,30.85"  # left,top,right,bottom (lon,lat,lon,lat)

# حدود الإسكندرية الموسّعة للتحقق من صحة النتيجة بعد الـ geocoding
# (أي نتيجة برا الحدود دي بتتتجاهل — غلطة من Nominatim)
ALEX_LAT_MIN, ALEX_LAT_MAX = 30.80, 31.40
ALEX_LNG_MIN, ALEX_LNG_MAX = 29.65, 30.30

def _is_in_alexandria(lat, lng):
    """تحقق إن الإحداثية دي داخل نطاق محافظة الإسكندرية."""
    return ALEX_LAT_MIN <= lat <= ALEX_LAT_MAX and ALEX_LNG_MIN <= lng <= ALEX_LNG_MAX

# مركز الإسكندرية تقريبًا — بيتستخدم كمرجع لفك أي Plus Code مختصر (short code) زي "6W4M+V4M"
# اللي محتاج نقطة قريبة معروفة عشان نكمّله لكود كامل ونطلع منه لات/لنج دقيق
ALEXANDRIA_REF_LAT, ALEXANDRIA_REF_LNG = 31.15, 29.95

_PLUS_CODE_PATTERN = re.compile(r'\b([A-Z0-9]{4,8}\+[A-Z0-9]{2,7})\b', re.IGNORECASE)

def try_decode_plus_code(raw_address):
    """
    لو العنوان فيه Plus Code (زي 6W4M+V4M)، بيفكه محليًا لإحداثيات lat/lng دقيقة
    من غير أي نداء لـ Nominatim خالص. أدق وأسرع بكتير من الـ geocoding النصي،
    خصوصًا إن الـ Plus Code ده بيوصف نقطة فعلية على الأرض مش اسم منطقة عامة.
    بيرجع (lat, lng) لو نجح، أو (None, None) لو مفيش plus code أو فشل الفك.
    """
    if not raw_address:
        return None, None
    m = _PLUS_CODE_PATTERN.search(raw_address.upper())
    if not m:
        return None, None
    code = m.group(1)
    try:
        if olc.isShort(code):
            full_code = olc.recoverNearest(code, ALEXANDRIA_REF_LAT, ALEXANDRIA_REF_LNG)
        else:
            full_code = code
        if not olc.isValid(full_code):
            return None, None
        decoded = olc.decode(full_code)
        return decoded.latitudeCenter, decoded.longitudeCenter
    except Exception as e:
        print(f"[plus_code] فشل فك '{code}': {type(e).__name__}: {e}")
        return None, None

# كلمات/أسماء شركات وفروع شائعة في نص العنوان مش بتفيد الـ geocoding وبتلخبطه
_BUSINESS_NOISE_PATTERN = re.compile(
    r'\b(?:amazon\s*now|amazon|qcd\d*\w*|rawahel|rwa7el|dsp|da\/dsp|talabat|breadfast|instashop)\b',
    re.IGNORECASE
)

# نمط كلمات دلالية على وحدة سكنية (شقة/عمارة/دور) بيتحذف هو والرقم اللي بعده
_UNIT_WORD_PATTERN = re.compile(r'\b(?:شقة|عمار[ةه]?|الدور|دور|apartment|apt\.?|floor|unit)\b\s*\d*', re.IGNORECASE)
# حروف ملاصقة لأرقام زي A1, B2 (رموز شقق/عمارات شائعة)
_LETTER_DIGIT_PATTERN = re.compile(r'\b[A-Za-z]\d+\b')
# كسور أرقام زي 3/102 (رقم دور/شقة)
_FRACTION_NUMBERS_PATTERN = re.compile(r'\b\d+/\d+\b')
# أي رقم منفرد باقي (63, 401, 4, 5...) كتوكن مستقل
_LONE_NUMBER_PATTERN = re.compile(r'(?<!\w)\d+(?!\w)')

def _strip_unit_numbers(text):
    """
    بيشيل أي رقم شقة/عمارة/دور من الجملة، من أي مكان فيها (مش بس آخرها)، وبيسيب
    الكلمات النصية (اسم الشارع/المنطقة) زي ما هي. مفيد للعناوين اللي مفيهاش فاصلة
    تفصل اسم الشارع عن تفاصيل الوحدة السكنية (زي "شارع النصر كمبيود اللوكس شقة A1 4 401 عمار 5").
    """
    t = _UNIT_WORD_PATTERN.sub('', text)
    t = _LETTER_DIGIT_PATTERN.sub('', t)
    t = _FRACTION_NUMBERS_PATTERN.sub('', t)
    t = _LONE_NUMBER_PATTERN.sub('', t)
    return re.sub(r'\s{2,}', ' ', t).strip(' ,-')

def _build_geocode_candidates(raw_address):
    """
    بيرجع لستة من نسخ مختلفة للعنوان (من الأخص للأعم)، عشان نجرب كل واحدة على Nominatim
    لحد ما نلاقي واحدة بترجع نتيجة. ده بديل التنظيف القديم اللي كان بيشيل معلومات
    مهمة من العنوان (زي رقم الشارع) فيخلي Nominatim مايعرفش يفهمه خالص.
    """
    address = raw_address

    # تحويل الأرقام العربية لإنجليزي
    arabic_to_english = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
    address = address.translate(arabic_to_english)

    # إزالة Plus Codes بس (زي 6W4M+V4M) لأن Nominatim مش بيدعمها خالص ومفيدش حتى كتلميح
    no_plus_code = re.sub(r'\b[A-Z0-9]{4,8}\+[A-Z0-9]{2,7}\b', '', address, flags=re.IGNORECASE).strip(' ,-')

    # إزالة أسماء شركات/فروع (زي "Amazon Now QCD2") لأنها مش جزء من العنوان الجغرافي فعليًا
    no_business = _BUSINESS_NOISE_PATTERN.sub('', no_plus_code).strip(' ,-')
    no_business = re.sub(r'\s{2,}', ' ', no_business).strip(' ,-')

    # إزالة كلمات الشقق/الأدوار/العيادات (وما بعدها) لو موجودة صراحة
    no_apt = re.sub(r'\b(?:apartment|apt|شقة|الدور|floor|door|دكتور|مستشفى|عيادة|clinic)\b.*$',
                     '', no_business, flags=re.IGNORECASE).strip(' ,-')

    # 🔴 تنظيف إضافي: شيل أي رقم شقة/عمارة/دور من أي مكان في الجملة (مش بس آخرها)
    # عشان عناوين زي "شارع النصر كمبيود اللوكس شقة A1 4 401 عمار 5" تبقى
    # "شارع النصر كمبيود اللوكس" بس — المطلوب مكان الشارع/المنطقة، مش دقة الشقة
    no_unit_numbers = _strip_unit_numbers(no_apt)

    lower_addr = no_business.lower()
    has_city = ('alexandria' in lower_addr or 'الإسكندرية' in no_business or 'اسكندرية' in no_business)
    is_arabic = bool(re.search(r'[\u0600-\u06FF]', no_business))
    city_suffix = ", الإسكندرية، مصر" if is_arabic else ", Alexandria, Egypt"

    candidates = []
    def add(c):
        c = (c or "").strip(' ,-')
        if c and c not in candidates:
            candidates.append(c)

    # 1. العنوان بعد شيل أرقام الشقة/العمارة المبعثرة (الأدق عادة — بيسيب اسم الشارع والكمبوند بس)
    add(no_unit_numbers if has_city else no_unit_numbers + city_suffix)
    # 2. العنوان بعد شيل اسم الشركة/الفرع + Plus Code + مدينة لو ناقصة، من غير قص أرقام الشقة
    add(no_business if has_city else no_business + city_suffix)
    # 3. نفس الحاجة بس من غير كلمات الشقة/الدور الصريحة
    add(no_apt if has_city else no_apt + city_suffix)
    # 4. اسم الشارع فقط (أول جزء قبل أول فاصلة) + المدينة — مفيد لو باقي العنوان فيه رقم عمارة/شقة مربك
    first_segment = re.split(r'[,\n]', no_apt)[0].strip()
    if first_segment:
        add(first_segment + city_suffix)
    # 5. لو العنوان فيه اسم كمبوند/مجمع (كلمة "كمبيود"/"كومباوند"/"compound") مش معروف لـ OSM غالبًا،
    # جرب بدونه تمامًا وسيب اسم الشارع فقط
    no_compound = re.sub(r'\b(?:كمبيود|كومباوند|كمبوند|compound)\b.*$', '', no_unit_numbers, flags=re.IGNORECASE).strip(' ,-')
    if no_compound and no_compound != no_unit_numbers:
        add(no_compound + (city_suffix if not has_city else ""))
    # 6. تحسين اسم منطقة معروف زي "Bab Shar" لو موجود
    if 'bab shar' in lower_addr:
        add(re.sub(r'(?i)bab shar', 'Bab Sharqi', first_segment) + city_suffix)
        add("Bab Sharqi, Alexandria, Egypt")
    if 'باب شرق' in no_business:
        add("باب شرق، الإسكندرية، مصر")
    # 7. أضعف احتمال: اسم المدينة بس (كملاذ أخير، هيدي نقطة تقريبية مش دقيقة)
    add("Alexandria, Egypt")

    return candidates


def geocode_address(address):
    """
    بيرجع (lat, lng) أو (None, None) لو العنوان مش واضح أو فشل الطلب.
    🔴 من غير كاش عادي — لكن بيشوف الأول resolved_locations (إحداثيات دقيقة اتجابت من
    جوجل مابس الفعلي عن طريق يوزرسكريبت GoogleMapsGeocode.txt)، لأنها أدق بكتير من أي
    geocoding نصي. لو مش موجودة، يجرب Plus Code محليًا، وبعدين Nominatim كملاذ أخير.
    """
    print(f"[geocode] called with address='{address}'")
    raw_address = (address or "").strip()
    if not raw_address:
        print("[geocode] empty address, skipping")
        return None, None

    # 🎯 أول أولوية: إحداثية دقيقة اتجابت من جوجل مابس الفعلي قبل كده لنفس العنوان بالظبط
    db = con()
    try:
        resolved = db.execute(
            "SELECT lat, lng FROM resolved_locations WHERE address=?", (raw_address,)
        ).fetchone()
    finally:
        db.close()
    if resolved and resolved["lat"] is not None:
        print(f"[geocode] Google Maps resolved location hit for '{raw_address}' -> {resolved['lat']}, {resolved['lng']}")
        return resolved["lat"], resolved["lng"]

    # 🎯 تاني أولوية: لو العنوان فيه Plus Code، فكّه محليًا فورًا بدون أي نداء لـ Nominatim.
    # ده أدق بكتير (بيحدد نقطة فعلية على الأرض) من أي geocoding نصي لاسم منطقة عامة.
    plus_lat, plus_lng = try_decode_plus_code(raw_address)
    if plus_lat is not None:
        print(f"[geocode] Plus Code decoded locally for '{raw_address}' -> {plus_lat}, {plus_lng}")
        return plus_lat, plus_lng

    try:
        candidates = _build_geocode_candidates(raw_address)
        print(f"[geocode] trying {len(candidates)} candidate(s) for '{raw_address}': {candidates}")

        for candidate in candidates:
            # احترام حد الطلبات (1 طلب/ثانية) بتاع Nominatim
            elapsed = _time.time() - _last_geocode_call[0]
            if elapsed < GEOCODE_MIN_INTERVAL:
                _time.sleep(GEOCODE_MIN_INTERVAL - elapsed)

            try:
                resp = requests.get(
                    "https://nominatim.openstreetmap.org/search",
                    params={
                        "q": candidate, "format": "json", "limit": 5,
                        "viewbox": ALEXANDRIA_VIEWBOX, "bounded": 1,
                    },
                    headers={"User-Agent": "RWA7EL-Dispatcher-Admin/1.0 (internal-tool)"},
                    timeout=8
                )
                _last_geocode_call[0] = _time.time()
                print(f"[geocode] candidate='{candidate}' status={resp.status_code} body_preview={resp.text[:150]!r}")
                results = resp.json() if resp.ok else []
            except Exception as e:
                print(f"[geocode] request EXCEPTION for candidate '{candidate}': {type(e).__name__}: {e}")
                results = []

            if results:
                # ناخد أفضل نتيجة حسب importance بدل أول نتيجة عشوائية
                # لكن مع فلتر: نشيل أي نتيجة برا حدود الإسكندرية (false positive من مدينة تانية)
                alex_results = [r for r in results if _is_in_alexandria(float(r["lat"]), float(r["lon"]))]
                if not alex_results:
                    print(f"[geocode] candidate='{candidate}' returned {len(results)} result(s) but ALL are outside Alexandria bounds, skipping")
                    continue
                best = max(alex_results, key=lambda r: float(r.get("importance", 0) or 0))
                lat, lng = float(best["lat"]), float(best["lon"])
                print(f"[geocode] SUCCESS for '{raw_address}' using candidate '{candidate}' -> {lat}, {lng} (importance={best.get('importance')})")
                return lat, lng

        # كل المحاولات فشلت — مش بنعمل unbounded fallback عشان منجيبش نتيجة غلط من مدينة تانية
        print(f"[geocode] all candidates failed (or returned results outside Alexandria) for '{raw_address}'")
        return None, None
    except Exception as e:
        import traceback
        print(f"[geocode] EXCEPTION for '{raw_address}': {type(e).__name__}: {e}")
        traceback.print_exc()
        return None, None

def haversine_km(lat1, lng1, lat2, lng2):
    """المسافة الفعلية بالكيلومتر بين نقطتين على سطح الأرض (خط مستقيم جوي، مش مسافة طريق فعلية)."""
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return 2 * R * math.asin(math.sqrt(a))

ADMIN_PASSWORD = "bankai"

@app.route("/")
@app.route("/admin")
def admin():
    pw = request.args.get("pw","")
    if pw != ADMIN_PASSWORD:
        return render_template("login.html"), 401
    return render_template("admin.html", public_url=public_url(), admin_pw=ADMIN_PASSWORD)

@app.route("/join")
def join():
    # Read as plain file to avoid Jinja2 parsing (join.html has no template vars)
    import os
    path = os.path.join(app.template_folder or 'templates', 'join.html')
    with open(path, 'r', encoding='utf-8') as f:
        html = f.read()
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}

@app.route("/rating")
def rating_page():
    pw = request.args.get("pw","")
    if pw != ADMIN_PASSWORD:
        return render_template("login.html"), 401
    return render_template("rawahel_driver_rating.html", admin_pw=ADMIN_PASSWORD)

@socketio.on('update_location')
def handle_location(data):
    name = data.get('name')
    lat = data.get('lat')
    lng = data.get('lng')
    accuracy = data.get('accuracy')
    heading = data.get('heading')
    speed = data.get('speed')
    altitude = data.get('altitude')
    ts = data.get('ts')
    if name and lat and lng:
        db = con()
        db.execute("UPDATE drivers SET lat=?, lng=? WHERE name=?", (lat, lng, name))
        db.commit()
        db.close()
        dist_m = data.get('dist_m')
        emit('location_updated', {
            'name': name, 'lat': lat, 'lng': lng,
            'accuracy': accuracy, 'heading': heading,
            'speed': speed, 'altitude': altitude,
            'dist_m': dist_m, 'ts': ts
        }, broadcast=True)

@socketio.on('request_battery_from_all')
def handle_request_battery():
    """الأدمن بيطلب من كل السواقين يبعتوا battery دلوقتي."""
    emit('send_battery_now', {}, broadcast=True)

@socketio.on('update_battery')
def handle_battery(data):
    name = data.get('name')
    level = data.get('level')   # 0-100 integer
    charging = data.get('charging', False)
    if name and level is not None:
        db = con()
        db.execute("UPDATE drivers SET battery=?, charging=? WHERE name=?", (int(level), 1 if charging else 0, name))
        db.commit()
        db.close()
        emit('battery_updated', {'name': name, 'level': int(level), 'charging': charging}, broadcast=True)

@app.route("/register", methods=["POST"])
def register():
    name = request.json.get("name", "").strip()
    if not name: return jsonify(ok=False, error="Name empty"), 400
    db = con()
    try:
        row = db.execute("SELECT * FROM drivers WHERE name=?", (name,)).fetchone()
        if row:
            if row["status"] in ("done", "break"):
                max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
                db.execute("UPDATE drivers SET status='waiting',created=?,out_time=NULL,done_time=NULL,break_start=NULL, break_paused_accum=0, break_pause_started=NULL,sort_order=? WHERE id=?",
                          (now(), max_order+1000, row["id"]))
                db.commit()
                db.close()
                db2 = con()
                pos = db2.execute("SELECT COUNT(*) as c FROM drivers WHERE status='waiting' AND sort_order <= (SELECT sort_order FROM drivers WHERE name=?)",(name,)).fetchone()["c"]
                db2.close()
                return jsonify(ok=True, position=pos)
        req_order = request.json.get('sort_order')
        if req_order is None:
            max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
            req_order = max_order+1000
        db.execute("INSERT INTO drivers(name,status,created,sort_order) VALUES(?,'waiting',?,?)", (name, now(), int(req_order)))
        db.commit()
        pos = db.execute("SELECT COUNT(*) as c FROM drivers WHERE status='waiting'").fetchone()["c"]
        db.close()
        emit_stats_all()
        return jsonify(ok=True, position=pos)
    except Exception as e:
        db.close()
        return jsonify(ok=False, error=str(e)), 500

def admin_auth():
    pw = request.headers.get("X-Admin-PW") or request.args.get("pw")
    return pw == ADMIN_PASSWORD

@app.route("/api/add_driver", methods=["POST"])
def add_driver():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    name = request.json.get("name", "").strip()
    if not name: return jsonify(ok=False, error="Name empty")
    db = con()
    try:
        req_order = request.json.get('sort_order')
        if req_order is None:
            max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
            req_order = max_order+1000
        db.execute("INSERT INTO drivers(name,status,created,sort_order) VALUES(?,'waiting',?,?)", (name, now(), int(req_order)))
        db.commit()
        emit_stats_all()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e))
    finally:
        db.close()

@app.route("/api/rename/<int:i>", methods=["POST"])
def rename_driver(i):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    new_name = request.json.get("name", "").strip()
    if not new_name: return jsonify(ok=False, error="Name empty")
    db = con()
    try:
        old_row = db.execute("SELECT name FROM drivers WHERE id=?", (i,)).fetchone()
        if old_row:
            old_name = old_row["name"]
            db.execute("UPDATE drivers SET name=? WHERE id=?", (new_name, i))
            db.execute("UPDATE driver_orders SET name=? WHERE name=?", (new_name, old_name))
            db.execute("UPDATE miss_history SET driver_name=? WHERE driver_name=?", (new_name, old_name))
            db.execute("UPDATE return_history SET driver_name=? WHERE driver_name=?", (new_name, old_name))
            db.execute("UPDATE done_history SET driver_name=? WHERE driver_name=?", (new_name, old_name))
            # old_name و new_name هنا هما الإيميل، فلازم نزامن جدولين التقييم كمان
            db.execute("UPDATE name_mapping SET email=? WHERE email=?", (new_name, old_name))
            db.execute("UPDATE driver_rating_notes SET email=? WHERE email=?", (new_name, old_name))
            db.commit()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e))
    finally:
        db.close()

@app.route("/api/list")
def lst():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = [dict(x) for x in db.execute("SELECT * FROM drivers ORDER BY sort_order, id")]
    db.close()
    return jsonify(rows)

@app.route("/api/orders_list")
def orders_list():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = [dict(x) for x in db.execute("SELECT name, orders, last_done FROM driver_orders ORDER BY orders DESC, last_done DESC")]
    db.close()
    return jsonify(rows)

@app.route("/api/dashboard_stats")
def dashboard_stats():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    t = today()
    # 🔴 الإحصائيات بقت مرتبطة بالشيفت الحالي (started_at) لو فيه شيفت شغال،
    # بدل ما تكون مرتبطة بتاريخ اليوم — عشان لو الشيفت عدّى نص الليل، الأرقام
    # متترجعش صفر لحد ما المستخدم يدوس "End Shift" فعليًا.
    shift_row = db.execute("SELECT started_at, started_at_display FROM shift_state WHERE id=1").fetchone()
    shift_active = bool(shift_row and shift_row["started_at"])
    if shift_active:
        since = shift_row["started_at"]
        del_row = db.execute("""SELECT COUNT(*) as c FROM done_history h
            WHERE h.timestamp>=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (since,)).fetchone()
        miss_row = db.execute("""SELECT COUNT(*) as c FROM miss_history h
            WHERE h.timestamp>=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (since,)).fetchone()
        ret_row = db.execute("""SELECT COUNT(*) as c FROM return_history h
            WHERE h.timestamp>=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (since,)).fetchone()
        label = shift_row["started_at_display"]
    else:
        del_row = db.execute("""SELECT COUNT(*) as c FROM done_history h
            WHERE h.date=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (t,)).fetchone()
        miss_row = db.execute("""SELECT COUNT(*) as c FROM miss_history h
            WHERE h.date=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (t,)).fetchone()
        ret_row = db.execute("""SELECT COUNT(*) as c FROM return_history h
            WHERE h.date=? AND h.driver_name IN (SELECT email FROM allowed_accounts)""", (t,)).fetchone()
        label = "Today"
    delivered = del_row["c"] if del_row and del_row["c"] else 0
    missed = miss_row["c"] if miss_row and miss_row["c"] else 0
    returned = ret_row["c"] if ret_row and ret_row["c"] else 0
    out_row = db.execute("SELECT COUNT(*) as c FROM drivers WHERE status='out'").fetchone()
    active = out_row["c"] if out_row and out_row["c"] else 0
    break_row = db.execute("SELECT COUNT(*) as c FROM drivers WHERE status='break'").fetchone()
    breaks = break_row["c"] if break_row and break_row["c"] else 0
    fines_row = db.execute("SELECT COALESCE(SUM(amount),0) as s FROM fines").fetchone()
    fines_total = fines_row["s"] if fines_row else 0
    db.close()
    total = delivered + missed + returned
    success_rate = round((delivered / total * 100), 1) if total > 0 else 0
    miss_rate = round((missed / total * 100), 1) if total > 0 else 0
    return_rate = round((returned / total * 100), 1) if total > 0 else 0
    return jsonify({
        "today_date": label,
        "shift_active": shift_active,
        "total_orders": total,
        "delivered": delivered,
        "missed": missed,
        "returned": returned,
        "active_drivers": active,
        "break_drivers": breaks,
        "success_rate": success_rate,
        "miss_rate": miss_rate,
        "return_rate": return_rate,
        "fines_total": fines_total
    })

@app.route("/api/public_url")
def api_public_url():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    return jsonify(url=public_url())

@app.route("/api/join_url")
def api_join_url():
    """Public endpoint — no auth — returns the join URL for QR generation."""
    u = public_url().rstrip("/") + "/join"
    return jsonify(url=u)

# ═══════════════════════════════════════════════════════════
# 📦 DISPATCHER ORDERS — الأوردرات الجايه من الديسباتشر (userscript)
# ═══════════════════════════════════════════════════════════
@app.route("/api/new_dispatcher_order", methods=["POST"])
def new_dispatcher_order():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    order_data = request.get_json(silent=True)
    if not order_data:
        return jsonify(error="No data"), 400
    order_hash = f"{order_data.get('pickup')}-{order_data.get('dropoff')}"
    db = con()
    try:
        existing = db.execute("SELECT id, picked_up_by FROM dispatcher_orders WHERE order_hash=?", (order_hash,)).fetchone()
        if existing:
            # لو الأوردر اتـ assign لسواق — تجاهله تماماً، متحدثوش ومتبعتش socket
            if existing["picked_up_by"] and existing["picked_up_by"].strip():
                return jsonify(ok=True, skipped=True)
            # مش assigned — تحديث عادي بدون مس picked_up_by
            db.execute("UPDATE dispatcher_orders SET time=?, distance=?, items=?, weight=? WHERE order_hash=?",
                      (order_data.get('time'), order_data.get('distance'), order_data.get('items'),
                       order_data.get('weight'), order_hash))
            db.commit()
            order_data['id'] = existing["id"]
            socketio.emit('dispatcher_order_time_updated', order_data)
        else:
            db.execute("""INSERT INTO dispatcher_orders(order_hash, time, distance, items, weight, pickup, dropoff, picked_up_by, created_at) 
                         VALUES(?,?,?,?,?,?,?,?,?)""",
                      (order_hash, order_data.get('time'), order_data.get('distance'), order_data.get('items'), 
                       order_data.get('weight'), order_data.get('pickup'), order_data.get('dropoff'),
                       order_data.get('pickedUpBy'), now()))
            db.commit()
            order_data['id'] = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            socketio.emit('incoming_dispatcher_order', order_data)
        socketio.emit('dispatcher_update')
    except Exception as e:
        db.close()
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()
    return jsonify(ok=True)

@app.route("/api/dispatcher_orders")
def get_dispatcher_orders():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = [dict(x) for x in db.execute("SELECT * FROM dispatcher_orders ORDER BY id DESC")]
    db.close()
    return jsonify(rows)

@app.route("/api/dispatcher_orders/delete/<int:i>", methods=["POST"])
def del_dispatcher_order(i):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM dispatcher_orders WHERE id=?", (i,))
    db.commit()
    db.close()
    socketio.emit('dispatcher_update')
    return jsonify(ok=True)

@app.route("/api/dispatcher_orders/clear", methods=["POST"])
def clear_dispatcher_orders():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM dispatcher_orders")
    db.commit()
    db.close()
    socketio.emit('dispatcher_update')
    return jsonify(ok=True)

@app.route("/api/dispatcher_orders/sync", methods=["POST"])
def sync_dispatcher_orders():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.get_json(silent=True) or {}
    live_hashes = data.get("hashes", [])
    if not isinstance(live_hashes, list):
        return jsonify(ok=False, error="hashes must be a list"), 400
    db = con()
    try:
        existing = db.execute("SELECT id, order_hash, picked_up_by FROM dispatcher_orders").fetchall()
        live_set = set(live_hashes)
        # الأوردرات الـ assigned (عندها picked_up_by) متتمسحش حتى لو اختفت من الديسباتشر
        removed_ids = [row["id"] for row in existing
                       if row["order_hash"] not in live_set
                       and not (row["picked_up_by"] and row["picked_up_by"].strip())]
        if removed_ids:
            db.executemany("DELETE FROM dispatcher_orders WHERE id=?", [(rid,) for rid in removed_ids])
            db.commit()
            socketio.emit('dispatcher_update')
        return jsonify(ok=True, removed=len(removed_ids))
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

@app.route("/api/dropoff_location", methods=["POST"])
def save_resolved_location():
    """
    بيستقبل إحداثية (lat/lng) اتجابت من يوزرسكريبت GoogleMapsGeocode.txt بعد ما فتح
    تاب جوجل مابس بعنوان معين وحل موقعه فعليًا. بنحفظها في resolved_locations، وهتتفضل
    ليها الأولوية على أي geocoding نصي (Nominatim) في endpoint الماب.
    """
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error="No data"), 400
    address = (data.get("address") or "").strip()
    lat = data.get("lat")
    lng = data.get("lng")
    if not address or lat is None or lng is None:
        return jsonify(ok=False, error="address, lat, lng required"), 400
    try:
        lat = float(lat)
        lng = float(lng)
    except (TypeError, ValueError):
        return jsonify(ok=False, error="lat/lng must be numeric"), 400
    db = con()
    try:
        db.execute(
            "INSERT OR REPLACE INTO resolved_locations(address, lat, lng, resolved_at) VALUES(?,?,?,?)",
            (address, lat, lng, now())
        )
        db.commit()
        print(f"[gmaps] تم حفظ إحداثية من جوجل مابس لـ '{address}' -> {lat}, {lng}")
        socketio.emit('dispatcher_update')
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

@app.route("/api/dropoff_location/check", methods=["POST"])
def check_resolved_location():
    """
    بيتحقق هل عندنا إحداثيات محفوظة من جوجل مابس لقايمة عناوين.
    بيقبل { addresses: ["عنوان1", "عنوان2", ...] }
    وبيرجع  { resolved: { "عنوان1": true, "عنوان2": false } }
    طلب واحد للـ batch بدل طلب لكل عنوان على حدة.
    """
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.get_json(silent=True)
    if not data:
        return jsonify(resolved={}), 400
    addresses = data.get("addresses") or []
    if not addresses or not isinstance(addresses, list):
        return jsonify(resolved={}), 400
    addresses = [str(a).strip() for a in addresses if a][:50]  # حد أقصى 50 عنوان في الطلب
    if not addresses:
        return jsonify(resolved={}), 400
    db = con()
    try:
        placeholders = ",".join("?" * len(addresses))
        rows = db.execute(
            f"SELECT address FROM resolved_locations WHERE address IN ({placeholders}) AND lat IS NOT NULL",
            addresses
        ).fetchall()
        found = {row["address"] for row in rows}
        result = {addr: (addr in found) for addr in addresses}
        return jsonify(resolved=result)
    finally:
        db.close()

@app.route("/api/dispatcher_orders/assign", methods=["POST"])
def assign_dispatcher_order():
    """
    يربط أوردر بسواق معين:
    - بيسجّل اسم السواق في picked_up_by على الأوردر
    - بيحوّل السواق من waiting → out تلقائياً
    - بيبعت socket event عشان الـ UI يتحدث فوراً
    """
    if not admin_auth():
        return jsonify(error="Unauthorized"), 401
    data = request.get_json(silent=True) or {}
    order_id   = data.get("order_id")
    driver_name = (data.get("driver_name") or "").strip()
    if not order_id or not driver_name:
        return jsonify(ok=False, error="order_id and driver_name required"), 400
    db = con()
    try:
        order = db.execute("SELECT * FROM dispatcher_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify(ok=False, error="Order not found"), 404
        driver = db.execute("SELECT * FROM drivers WHERE name=?", (driver_name,)).fetchone()
        if not driver:
            return jsonify(ok=False, error="Driver not found"), 404
        if driver["status"] != "waiting":
            return jsonify(ok=False, error=f"Driver is {driver['status']}, not waiting"), 409
        t = now()
        db.execute("UPDATE dispatcher_orders SET picked_up_by=? WHERE id=?", (driver_name, order_id))
        db.execute("UPDATE drivers SET status='out', out_time=? WHERE name=?", (t, driver_name))
        db.commit()
        socketio.emit("dispatcher_update")
        socketio.emit("driver_assigned", {
            "order_id": order_id,
            "driver_name": driver_name,
            "pickup": order["pickup"],
            "dropoff": order["dropoff"]
        })
        return jsonify(ok=True, order_id=order_id, driver_name=driver_name)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

# ═══════════════════════════════════════════════════════════
# 🗺️ ORDERS MAP DATA — إحداثيات كل أوردر مفتوح لعرضها على خريطة
# ═══════════════════════════════════════════════════════════
@app.route("/api/dispatcher_orders/map")
def dispatcher_orders_map():
    """
    خريطة الأوردرات — بيرجع إحداثيات (lat/lng) كل أوردر مفتوح (dropoff فقط).
    🔴 من غير كاش خالص: كل نداء بيعمل geocode حي (live) من الصفر لكل عنوان تسليم.
    لو ?scope=all اتبعتت، بيرجع كل الأوردرات (حتى المُعيَّنة لسائق) مش بس اللي لسه من غير سائق.
    """
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    scope_all = request.args.get("scope") == "all"
    db = con()
    orders = [dict(x) for x in db.execute(
        "SELECT id, pickup, dropoff, picked_up_by, items, weight, distance FROM dispatcher_orders ORDER BY id DESC"
    )]
    db.close()
    target_orders = orders if scope_all else [
        o for o in orders if not (o.get("picked_up_by") and str(o["picked_up_by"]).strip())
    ]
    print(f"[map] scope={'all' if scope_all else 'unassigned'} orders to geocode: {len(target_orders)}")
    points = []
    for o in target_orders:
        d_lat, d_lng = geocode_address(o["dropoff"])
        assigned_to = (o.get("picked_up_by") or "").strip()
        points.append({
            "id": o["id"], "pickup": o["pickup"], "dropoff": o["dropoff"],
            "dropoff_lat": d_lat, "dropoff_lng": d_lng,
            "dropoff_failed": d_lat is None,
            "items": o.get("items"), "weight": o.get("weight"), "distance": o.get("distance"),
            "assigned": bool(assigned_to), "assigned_to": assigned_to,
        })
    ok_count = sum(1 for p in points if p["dropoff_lat"] is not None)
    print(f"[map] geocoded successfully: {ok_count}/{len(points)} dropoffs")
    return jsonify(points=points)

@app.route("/api/allowed_accounts")
def list_allowed_accounts():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = db.execute("SELECT email FROM allowed_accounts ORDER BY email").fetchall()
    db.close()
    return jsonify([r["email"] for r in rows])

@app.route("/api/allowed_accounts/add", methods=["POST"])
def add_allowed_account():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    email = (request.json.get("email", "") or "").strip().lower()
    if not email: return jsonify(ok=False, error="Email empty")
    db = con()
    db.execute("INSERT OR IGNORE INTO allowed_accounts(email) VALUES(?)", (email,))
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/allowed_accounts/remove", methods=["POST"])
def remove_allowed_account():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    email = (request.json.get("email", "") or "").strip().lower()
    db = con()
    db.execute("DELETE FROM allowed_accounts WHERE email=?", (email,))
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/allowed_accounts/clear", methods=["POST"])
def clear_allowed_accounts():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM allowed_accounts")
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/allowed_accounts/import_xlsx", methods=["POST"])
def import_allowed_accounts_xlsx():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    if "file" not in request.files:
        return jsonify(ok=False, error="No file uploaded"), 400
    f = request.files["file"]
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        emails = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell is None: continue
                val = str(cell).strip().lower()
                if "@" in val and val != "email address":
                    emails.append(val)
                    break
        emails = list(set(emails))
    except Exception as e:
        return jsonify(ok=False, error=f"Failed to read file: {e}"), 400
    if not emails:
        return jsonify(ok=False, error="No valid emails found in file"), 400
    db = con()
    db.executemany("INSERT OR IGNORE INTO allowed_accounts(email) VALUES(?)", [(e,) for e in emails])
    db.commit()
    total = db.execute("SELECT COUNT(*) as c FROM allowed_accounts").fetchone()["c"]
    db.close()
    return jsonify(ok=True, imported=len(emails), total=total)

# ============================================================
# 🔗 name_mapping — ربط إيميل المندوب (المستخدم في نظام الكيو) بالاسم الحقيقي
# (المستخدم في نظام التقييم)
# ============================================================
@app.route("/api/name_mapping")
def list_name_mapping():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = db.execute("SELECT email, full_name FROM name_mapping ORDER BY full_name").fetchall()
    db.close()
    return jsonify([{"email": r["email"], "name": r["full_name"]} for r in rows])

@app.route("/api/name_mapping/add", methods=["POST"])
def add_name_mapping():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    email = (request.json.get("email", "") or "").strip().lower()
    name = (request.json.get("name", "") or "").strip()
    if not email or "@" not in email or not name:
        return jsonify(ok=False, error="email/name غير صحيحين")
    db = con()
    db.execute("INSERT INTO name_mapping(email, full_name) VALUES(?,?) ON CONFLICT(email) DO UPDATE SET full_name=excluded.full_name", (email, name))
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/name_mapping/remove", methods=["POST"])
def remove_name_mapping():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    email = (request.json.get("email", "") or "").strip().lower()
    db = con()
    db.execute("DELETE FROM name_mapping WHERE email=?", (email,))
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/name_mapping/clear", methods=["POST"])
def clear_name_mapping():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM name_mapping")
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/name_mapping/import_xlsx", methods=["POST"])
def import_name_mapping_xlsx():
    """يستورد إكسل: العمود الأول = الاسم، العمود التاني = الإيميل.
    نفس فكرة استيراد allowed_accounts بالظبط، بس بعمودين."""
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    if "file" not in request.files:
        return jsonify(ok=False, error="No file uploaded"), 400
    f = request.files["file"]
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        pairs = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            name_cell, email_cell = row[0], row[1]
            if not name_cell or not email_cell:
                continue
            name = str(name_cell).strip()
            email = str(email_cell).strip().lower()
            if "@" not in email or name.lower() in ("name", "الاسم"):
                continue
            pairs.append((email, name))
    except Exception as e:
        return jsonify(ok=False, error=f"Failed to read file: {e}"), 400
    if not pairs:
        return jsonify(ok=False, error="No valid name/email pairs found in file"), 400
    db = con()
    db.executemany(
        "INSERT INTO name_mapping(email, full_name) VALUES(?,?) ON CONFLICT(email) DO UPDATE SET full_name=excluded.full_name",
        pairs
    )
    db.commit()
    total = db.execute("SELECT COUNT(*) as c FROM name_mapping").fetchone()["c"]
    db.close()
    return jsonify(ok=True, imported=len(pairs), total=total)

@app.route("/api/export_misses")
def export_misses():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = db.execute("""
        SELECT d.name, d.misses, h.order_id, h.reason, h.timestamp
        FROM drivers d
        LEFT JOIN miss_history h ON d.name = h.driver_name
        WHERE d.misses > 0
        ORDER BY d.misses DESC, d.name ASC, h.timestamp DESC
    """).fetchall()
    db.close()
    si = io.StringIO()
    si.write('\ufeff')
    cw = csv.writer(si)
    cw.writerow(['Driver Name', 'Total Misses', 'Order ID', 'Miss Reason', 'Time and Date'])
    for r in rows:
        cw.writerow([r['name'], r['misses'], r['order_id'] or 'N/A', r['reason'] or 'N/A', r['timestamp'] or 'N/A'])
    return Response(si.getvalue(), mimetype="text/csv")

@app.route("/api/export_returns")
def export_returns():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = db.execute("""
        SELECT d.name, d.returns, h.order_id, h.timestamp
        FROM drivers d
        LEFT JOIN return_history h ON d.name = h.driver_name
        WHERE d.returns > 0
        ORDER BY d.returns DESC, d.name ASC, h.timestamp DESC
    """).fetchall()
    db.close()
    si = io.StringIO()
    si.write('\ufeff')
    cw = csv.writer(si)
    cw.writerow(['Driver Name', 'Total Returns', 'Order ID', 'Time and Date'])
    for r in rows:
        cw.writerow([r['name'], r['returns'], r['order_id'] or 'N/A', r['timestamp'] or 'N/A'])
    return Response(si.getvalue(), mimetype="text/csv")

# ============================================================
# 💰 الغرامات (Fines) — إدارة كاملة + Export
# ============================================================
@app.route("/api/fines")
def list_fines():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = [dict(x) for x in db.execute("SELECT * FROM fines ORDER BY timestamp DESC")]
    total_amount = sum(r["amount"] or 0 for r in rows)
    db.close()
    return jsonify(fines=rows, total=total_amount, count=len(rows))

@app.route("/api/fines/add", methods=["POST"])
def add_fine():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.json or {}
    driver_name = (data.get("driver_name", "") or "").strip()
    reason = (data.get("reason", "") or "").strip()
    try:
        amount = float(data.get("amount"))
    except (TypeError, ValueError):
        return jsonify(ok=False, error="المبلغ غير صحيح"), 400
    if not driver_name:
        return jsonify(ok=False, error="اسم السواق مطلوب"), 400
    if amount <= 0:
        return jsonify(ok=False, error="المبلغ لازم يكون أكبر من صفر"), 400
    db = con()
    db.execute("INSERT INTO fines(driver_name, amount, reason, timestamp, date) VALUES(?,?,?,?,?)",
              (driver_name, amount, reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
    db.close()
    socketio.emit("fines_updated", {})
    return jsonify(ok=True, id=new_id)

@app.route("/api/fines/delete/<int:i>", methods=["POST"])
def delete_fine(i):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM fines WHERE id=?", (i,))
    db.commit()
    db.close()
    socketio.emit("fines_updated", {})
    return jsonify(ok=True)

@app.route("/api/fines/edit/<int:i>", methods=["POST"])
def edit_fine(i):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.json or {}
    db = con()
    row = db.execute("SELECT * FROM fines WHERE id=?", (i,)).fetchone()
    if not row:
        db.close()
        return jsonify(ok=False, error="الغرامة غير موجودة"), 404
    driver_name = (data.get("driver_name", row["driver_name"]) or "").strip()
    reason = data.get("reason", row["reason"])
    try:
        amount = float(data.get("amount", row["amount"]))
    except (TypeError, ValueError):
        db.close()
        return jsonify(ok=False, error="المبلغ غير صحيح"), 400
    db.execute("UPDATE fines SET driver_name=?, amount=?, reason=? WHERE id=?",
              (driver_name, amount, reason, i))
    db.commit()
    db.close()
    socketio.emit("fines_updated", {})
    return jsonify(ok=True)

@app.route("/api/fines/clear", methods=["POST"])
def clear_fines():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM fines")
    db.commit()
    db.close()
    socketio.emit("fines_updated", {})
    return jsonify(ok=True)

@app.route("/api/fines/export")
def export_fines():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    rows = db.execute("SELECT * FROM fines ORDER BY driver_name ASC, timestamp DESC").fetchall()
    db.close()
    NAVY = "0A192F"
    NAVY2 = "112240"
    RED = "DC2626"
    RED_BG = "FEF2F2"
    WHITE = "FFFFFF"
    GREY = "F1F5F9"
    BORDER_CLR = "E2E8F0"
    TEXT = "1E293B"
    MUTED = "64748B"
    FONT_NAME = "Calibri"
    f_title = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
    f_subtitle = Font(name=FONT_NAME, size=10, color="CBD5E1")
    f_header = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    f_normal = Font(name=FONT_NAME, size=10.5, color=TEXT)
    f_amount = Font(name=FONT_NAME, size=10.5, bold=True, color=RED)
    f_muted = Font(name=FONT_NAME, size=9.5, italic=True, color=MUTED)
    f_kpi_label = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    f_kpi_value = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
    fill_navy = PatternFill("solid", fgColor=NAVY)
    fill_navy2 = PatternFill("solid", fgColor=NAVY2)
    fill_red = PatternFill("solid", fgColor=RED)
    fill_white = PatternFill("solid", fgColor=WHITE)
    fill_grey = PatternFill("solid", fgColor=GREY)
    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center", indent=1)
    def zebra(r):
        return fill_grey if r % 2 == 0 else fill_white
    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    def setup_print(ws):
        ws.page_setup.orientation = "landscape"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True
    wb = openpyxl.Workbook()
    # ── Sheet 1: كل الغرامات (تفاصيل) ──
    ws = wb.active
    ws.title = "Fines"
    ws.sheet_properties.tabColor = RED
    ws.sheet_view.showGridLines = False
    setup_print(ws)
    total_amount = sum(r["amount"] or 0 for r in rows)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value=f"💰  Fines Report — {len(rows)} fines")
    c.font = f_title
    c.fill = fill_navy
    c.alignment = center
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1, value=f"Generated {now_str}")
    c.font = f_subtitle
    c.fill = fill_navy
    c.alignment = center
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A4:B4")
    c = ws.cell(row=4, column=1, value="TOTAL FINES")
    c.font = f_kpi_label
    c.fill = fill_red
    c.alignment = center
    ws.merge_cells("A5:B5")
    c = ws.cell(row=5, column=1, value=total_amount)
    c.font = f_kpi_value
    c.fill = fill_red
    c.alignment = center
    c.number_format = "#,##0.00"
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 32
    r = 7
    headers = ["Driver Name", "Amount", "Reason", "Date / Time"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = f_header
        cell.fill = fill_navy2
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    r += 1
    fill_red_bg = PatternFill("solid", fgColor=RED_BG)
    for row in rows:
        fill = fill_red_bg if r % 2 == 0 else fill_white
        vals = [row["driver_name"], row["amount"], row["reason"] or "N/A", row["timestamp"]]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col in (2, 4) else left_a
            if col == 2:
                cell.font = f_amount
                cell.number_format = "#,##0.00"
            else:
                cell.font = f_normal
        r += 1
    if not rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value="لا توجد غرامات مسجلة")
        c.font = f_muted
        c.alignment = center
    autosize(ws, [28, 14, 36, 22])
    # ── Sheet 2: حسب السواق (إجمالي لكل سواق) ──
    ws2 = wb.create_sheet("Per Driver")
    ws2.sheet_properties.tabColor = NAVY
    ws2.sheet_view.showGridLines = False
    setup_print(ws2)
    per_driver = {}
    for row in rows:
        nm = row["driver_name"]
        per_driver.setdefault(nm, {"count": 0, "total": 0.0})
        per_driver[nm]["count"] += 1
        per_driver[nm]["total"] += row["amount"] or 0
    ws2.merge_cells("A1:C1")
    c = ws2.cell(row=1, column=1, value="📋  Fines Per Driver")
    c.font = f_title
    c.fill = fill_navy
    c.alignment = center
    ws2.row_dimensions[1].height = 32
    r2 = 3
    for i, h in enumerate(["Driver Name", "Fines Count", "Total Amount"], start=1):
        cell = ws2.cell(row=r2, column=i, value=h)
        cell.font = f_header
        cell.fill = fill_navy2
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[r2].height = 22
    ws2.freeze_panes = ws2.cell(row=r2 + 1, column=1)
    r2 += 1
    for nm in sorted(per_driver, key=lambda k: per_driver[k]["total"], reverse=True):
        d = per_driver[nm]
        fill = zebra(r2)
        vals = [nm, d["count"], d["total"]]
        for col, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r2, column=col, value=v)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col > 1 else left_a
            if col == 3:
                cell.font = f_amount
                cell.number_format = "#,##0.00"
            else:
                cell.font = f_normal
        r2 += 1
    if not per_driver:
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
        c = ws2.cell(row=r2, column=1, value="لا توجد غرامات مسجلة")
        c.font = f_muted
        c.alignment = center
    autosize(ws2, [28, 16, 18])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    pad = lambda n: str(n).zfill(2)
    dt = datetime.now()
    fname = f"Fines_Report_{dt.year}-{pad(dt.month)}-{pad(dt.day)}_{pad(dt.hour)}-{pad(dt.minute)}.xlsx"
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     headers={'Content-Disposition': f'attachment;filename={fname}'})

# ============================================================
# ⭐ نظام التقييم (Rating) — بيانات لايف من قاعدة بيانات الكيو
# name في كل الجداول التحتيه = الإيميل، وهنا بنحوله للاسم الحقيقي
# عن طريق name_mapping قبل ما نرجعه لصفحة التقييم
# ============================================================
@app.route("/api/rating_data")
def rating_data():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    mapping_rows = db.execute("SELECT email, full_name FROM name_mapping").fetchall()
    email_to_name = {r["email"]: r["full_name"] for r in mapping_rows}
    drivers_rows = db.execute("SELECT name, misses FROM drivers").fetchall()
    done_rows = db.execute(
        "SELECT driver_name, timestamp, date FROM done_history ORDER BY timestamp ASC"
    ).fetchall()
    miss_rows = db.execute(
        "SELECT driver_name, order_id, reason, timestamp, date FROM miss_history ORDER BY timestamp ASC"
    ).fetchall()
    notes_rows = db.execute("SELECT email, notes FROM driver_rating_notes").fetchall()
    notes_by_email = {r["email"]: r["notes"] for r in notes_rows}
    # كل البريدات اللي ليها أي وجود (drivers أو mapping) عشان حتى مين لسه معندوش أوردرات يظهر
    all_emails = set(email_to_name.keys()) | set(r["name"] for r in drivers_rows)
    result = {}
    for email in all_emails:
        result[email] = {
            "email": email,
            "name": email_to_name.get(email, email),  # لو مفيش mapping، نعرض الإيميل نفسه
            "orders": [],     # [{date, time}]
            "misses": [],     # [{order_id, reason, date, time}]
            "notes": notes_by_email.get(email, "") or ""
        }
    for r in done_rows:
        e = r["driver_name"]
        if e not in result:
            continue
        result[e]["orders"].append({"date": r["date"], "time": r["timestamp"]})
    for r in miss_rows:
        e = r["driver_name"]
        if e not in result:
            continue
        result[e]["misses"].append({
            "order_id": r["order_id"] or "—",
            "reason": r["reason"] or "بدون سبب مسجل",
            "date": r["date"],
            "time": r["timestamp"]
        })
    db.close()
    out = []
    for email, d in result.items():
        d["orders_count"] = len(d["orders"])
        d["misses_count"] = len(d["misses"])
        out.append(d)
    out.sort(key=lambda x: x["name"])
    return jsonify(out)

@app.route("/api/rating_notes/save", methods=["POST"])
def save_rating_notes():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    data = request.json or {}
    email = (data.get("email", "") or "").strip().lower()
    notes = (data.get("notes", "") or "").strip()
    if not email:
        return jsonify(ok=False, error="email مطلوب"), 400
    db = con()
    db.execute(
        "INSERT INTO driver_rating_notes(email, notes, updated_at) VALUES(?,?,?) "
        "ON CONFLICT(email) DO UPDATE SET notes=excluded.notes, updated_at=excluded.updated_at",
        (email, notes, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    db.commit()
    db.close()
    return jsonify(ok=True)

@app.route("/api/reorder", methods=["POST"])
def reorder():
    ids = request.json.get("ids", [])
    if not ids: return jsonify(ok=False, error="No ids"), 400
    db = con()
    try:
        for i, driver_id in enumerate(ids):
            db.execute("UPDATE drivers SET sort_order=? WHERE id=?", ((i+1)*10, driver_id))
        db.commit()
        db.close()
        return jsonify(ok=True)
    except Exception as e:
        db.close()
        return jsonify(ok=False, error=str(e)), 500

@app.route("/api/status/<int:i>/<s>", methods=["POST"])
def st(i, s):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    t = now()
    affected_name = None
    if s == "out":
        db.execute("UPDATE drivers SET status=?, out_time=?, break_start=NULL, break_paused_accum=0, break_pause_started=NULL WHERE id=?", (s, t, i))
    elif s == "miss":
        req_data = request.get_json(silent=True) or {}
        order_id = req_data.get("order_id", "")
        reason = req_data.get("reason", "")
        row = db.execute("SELECT name FROM drivers WHERE id=?", (i,)).fetchone()
        if row:
            affected_name = row["name"]
            db.execute("INSERT INTO miss_history(driver_name, order_id, reason, timestamp, date) VALUES(?,?,?,?,?)",
                      (row["name"], order_id, reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
            max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
            db.execute("UPDATE drivers SET status='waiting', created=?, out_time=NULL, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=?, misses=misses+1 WHERE id=?",
                      (t, max_order+1000, i))
            socketio.emit("rating_updated", {"email": row["name"]})
    elif s == "return":
        req_data = request.get_json(silent=True) or {}
        order_id = req_data.get("order_id", "")
        row = db.execute("SELECT name FROM drivers WHERE id=?", (i,)).fetchone()
        if row:
            affected_name = row["name"]
            db.execute("INSERT INTO return_history(driver_name, order_id, timestamp, date) VALUES(?,?,?,?)",
                      (row["name"], order_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
            max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
            db.execute("UPDATE drivers SET status='waiting', created=?, out_time=NULL, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=?, returns=returns+1 WHERE id=?",
                      (t, max_order+1000, i))
    elif s == "break":
        db.execute("UPDATE drivers SET status=?, break_start=?, break_paused=0, break_paused_accum=0, break_pause_started=NULL WHERE id=?", (s, t, i))
    elif s == "break_pause":
        db.execute("UPDATE drivers SET break_paused=1, break_pause_started=? WHERE id=?", (t, i))
    elif s == "break_resume":
        row = db.execute("SELECT break_pause_started, break_paused_accum FROM drivers WHERE id=?", (i,)).fetchone()
        added = 0
        if row and row["break_pause_started"]:
            try:
                paused_at = datetime.strptime(row["break_pause_started"], "%H:%M:%S")
                resumed_at = datetime.strptime(t, "%H:%M:%S")
                if resumed_at < paused_at:
                    resumed_at += timedelta(days=1)
                added = int((resumed_at - paused_at).total_seconds())
            except Exception:
                added = 0
        accum = (row["break_paused_accum"] or 0) if row else 0
        db.execute("UPDATE drivers SET break_paused=0, break_paused_accum=?, break_pause_started=NULL WHERE id=?",
                  (accum + added, i))
    elif s == "break_out":
        max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
        db.execute("UPDATE drivers SET status='waiting', created=?, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=? WHERE id=?",
                  (t, max_order+1000, i))
    elif s == "done":
        row = db.execute("SELECT name FROM drivers WHERE id=?", (i,)).fetchone()
        if row:
            name = row["name"]
            affected_name = name
            existing = db.execute("SELECT orders FROM driver_orders WHERE name=?", (name,)).fetchone()
            if existing:
                db.execute("UPDATE driver_orders SET orders=orders+1, last_done=? WHERE name=?", (t, name))
            else:
                db.execute("INSERT INTO driver_orders(name, orders, last_done) VALUES(?,1,?)", (name, t))
            db.execute("INSERT INTO done_history(driver_name, timestamp, date) VALUES(?,?,?)",
                      (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
            max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
            db.execute("UPDATE drivers SET status='waiting', created=?, out_time=NULL, done_time=?, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=? WHERE id=?",
                      (t, t, max_order+1000, i))
            socketio.emit("rating_updated", {"email": name})
    else:
        db.execute("UPDATE drivers SET status=? WHERE id=?", (s, i))
    db.commit()
    db.close()
    if affected_name:
        emit_stats_all()
    return jsonify(ok=True)

@app.route("/api/delete/<int:i>", methods=["POST"])
def delete_driver(i):
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    # جيب اسم الطيار قبل الحذف عشان نبعتله force_logout
    row = db.execute("SELECT name FROM drivers WHERE id=?", (i,)).fetchone()
    if row:
        socketio.emit("force_logout", {"name": row["name"]})
    db.execute("DELETE FROM drivers WHERE id=?", (i,))
    db.commit()
    db.close()
    emit_stats_all()
    return jsonify(ok=True)

@app.route("/api/reset_orders", methods=["POST"])
def reset_orders():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    db.execute("DELETE FROM driver_orders")
    db.commit()
    db.close()
    emit_stats_all()
    return jsonify(ok=True)

@app.route("/api/reset_dashboard", methods=["POST"])
def reset_dashboard():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    try:
        db.execute("DELETE FROM done_history")
        db.execute("DELETE FROM miss_history")
        db.execute("DELETE FROM return_history")
        db.execute("UPDATE drivers SET misses=0, returns=0")
        db.commit()
    except Exception as e:
        db.close()
        return jsonify(ok=False, error=str(e)), 500
    db.close()
    emit_stats_all()
    return jsonify(ok=True)

@app.route("/api/self_return/<name>", methods=["POST"])
def self_return(name):
    """Driver-initiated return to waiting — no admin auth required.
    If driver was 'out', it counts as a return order (logged in return_history)."""
    db = con()
    row = db.execute("SELECT id, status FROM drivers WHERE name=?", (name,)).fetchone()
    if not row:
        db.close()
        return jsonify(ok=False, error="Driver not found")
    if row["status"] not in ("out", "break"):
        db.close()
        return jsonify(ok=False, error="Not in out/break status")
    t = now()
    was_out = row["status"] == "out"
    if was_out:
        # سجّل كـ order done في التاريخ وزوّد عداد الأوردرات
        existing = db.execute("SELECT orders FROM driver_orders WHERE name=?", (name,)).fetchone()
        if existing:
            db.execute("UPDATE driver_orders SET orders=orders+1, last_done=? WHERE name=?", (t, name))
        else:
            db.execute("INSERT INTO driver_orders(name, orders, last_done) VALUES(?,1,?)", (name, t))
        db.execute("INSERT INTO done_history(driver_name, timestamp, date) VALUES(?,?,?)",
                  (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
        socketio.emit("rating_updated", {"email": name})
    max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
    db.execute(
        "UPDATE drivers SET status='waiting', created=?, out_time=NULL, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=? WHERE id=?",
        (t, max_order + 1000, row["id"])
    )
    db.commit()
    db.close()
    # لو كان out وعمل done → الـ rank بيأثر على الكل، ابعت للكل
    # لو كان break بس → مش محتاج تحديث stats خالص
    if was_out:
        emit_stats_all()
    return jsonify(ok=True)

@app.route("/api/position/<name>")
def get_position(name):
    db = con()
    row = db.execute("SELECT * FROM drivers WHERE name=?", (name,)).fetchone()
    if not row:
        db.close()
        return jsonify(found=False)
    pos = None
    if row["status"] == "waiting":
        pos = db.execute("SELECT COUNT(*) as c FROM drivers WHERE status='waiting' AND sort_order <=?", (row["sort_order"],)).fetchone()["c"]
    elapsed = 0
    if row["status"] == "break" and row["break_start"]:
        try:
            fmt = "%H:%M:%S"
            start_dt = datetime.strptime(row["break_start"], fmt)
            now_dt = datetime.strptime(now(), fmt)
            diff = (now_dt - start_dt).total_seconds()
            if diff < 0: diff += 24 * 3600
            elapsed = int(diff)
        except:
            pass
    db.close()
    return jsonify(found=True, status=row["status"], position=pos, break_start=row["break_start"], break_paused=row["break_paused"], elapsed=elapsed)

def compute_stats(db, name):
    orders_row = db.execute("SELECT orders FROM driver_orders WHERE name=?", (name,)).fetchone()
    orders = orders_row["orders"] if orders_row else 0
    driver_row = db.execute(
        "SELECT misses, returns, battery, charging, lat, lng FROM drivers WHERE name=?", (name,)
    ).fetchone()
    misses   = driver_row["misses"]   if driver_row else 0
    returns  = driver_row["returns"]  if driver_row else 0
    battery  = driver_row["battery"]  if driver_row else None
    charging = bool(driver_row["charging"]) if driver_row else False
    lat      = driver_row["lat"]      if driver_row else None
    lng      = driver_row["lng"]      if driver_row else None
    def score(o, m): return o * 10 - m * 25
    my_score = score(orders, misses)
    all_drivers = db.execute("""
        SELECT d.name, d.misses, d.returns, COALESCE(do.orders,0) as orders
        FROM drivers d
        LEFT JOIN driver_orders do ON d.name = do.name
    """).fetchall()
    total = len(all_drivers)
    rank = 1
    for row in all_drivers:
        if row["name"] == name:
            continue
        if score(row["orders"], row["misses"]) > my_score:
            rank += 1
    return {
        "orders": orders, "misses": misses, "returns": returns,
        "rank": rank, "total": total,
        "battery": battery, "charging": charging,
        "lat": lat, "lng": lng
    }

def send_position_push(name, position):
    """
    بيبعت Push Notification للطيار حسب ترتيبه في الانتظار:
      - ترتيب 1 → your_turn  (أقوى تنبيه)
      - ترتيب 2 أو 3 → almost (تنبيه تحذيري)
      - أكبر من 3 → مش بنبعت إشعار
    بيشتغل في thread منفصل عشان ميبطّلش الـ response.
    """
    if not PUSH_ENABLED:
        return
    if position is None or position > 3:
        return

    import threading

    def _do_push():
        db = con()
        try:
            row = db.execute(
                "SELECT subscription FROM push_subscriptions WHERE name=?", (name,)
            ).fetchone()
            if not row:
                return

            sub = json.loads(row["subscription"])

            if position == 1:
                notif_type = "your_turn"
                title = "🚨 دورك دلوقتي!"
                body  = f"يا {name} — الأدمن هيطلبك! استعد فوراً."
            elif position == 2:
                notif_type = "almost"
                title = "⚡ تاني واحد — استعد!"
                body  = f"يا {name} — ترتيبك #2، واحد قبلك بس."
            elif position == 3:
                notif_type = "almost"
                title = "📋 ترتيبك #3 — جهّز نفسك"
                body  = f"يا {name} — اتنين قبلك بس."

            payload = json.dumps({
                "type":  notif_type,
                "title": title,
                "body":  body,
                "url":   "/join"
            }, ensure_ascii=False)

            def _send_once():
                webpush(
                    subscription_info=sub,
                    data=payload,
                    vapid_private_key=VAPID_PRIVATE_KEY_PATH,
                    vapid_claims=VAPID_CLAIMS
                )

            # بعت الإشعار 3 مرات بفاصل 30 ثانية — عشان لو مردّش يجيه تاني
            import time as _time
            for attempt in range(3):
                try:
                    _send_once()
                    print(f"[push] ✅ sent {notif_type} → {name} (pos #{position}) attempt #{attempt+1}")
                except WebPushException as e:
                    if hasattr(e, "response") and e.response and e.response.status_code in (404, 410):
                        db.execute("DELETE FROM push_subscriptions WHERE name=?", (name,))
                        db.commit()
                        print(f"[push] ❌ subscription expired for {name} — deleted")
                        break
                    print(f"[push] ❌ WebPushException for {name}: {e}")
                    break
                except Exception as e:
                    print(f"[push] ❌ Error for {name}: {e}")
                    break
                # لو مش آخر محاولة — استنى 30 ثانية
                if attempt < 2:
                    _time.sleep(30)

        except Exception as e:
            print(f"[push] ❌ outer error for {name}: {e}")
        finally:
            db.close()

    threading.Thread(target=_do_push, daemon=True).start()


def emit_stats_all(only_name=None):
    """بيبعت stats_updated socket.
    لو only_name اتبعت، بيبعت لسواق واحد بس (أكفأ بكتير).
    لو None، بيبعت لكل السواقين (للحالات اللي بيأثر فيها ترتيب الكل زي miss/done).
    + بيبعت position push تلقائياً لأي طيار في المراكز 1/2/3
    """
    db = con()
    try:
        if only_name:
            names = [only_name]
        else:
            names = [r["name"] for r in db.execute("SELECT name FROM drivers").fetchall()]
        for name in names:
            try:
                stats = compute_stats(db, name)
                stats["name"] = name
                socketio.emit("stats_updated", stats)

                # ── Smart Position Push ──
                row = db.execute(
                    "SELECT status, sort_order FROM drivers WHERE name=?", (name,)
                ).fetchone()
                if row and row["status"] == "waiting":
                    pos = db.execute(
                        "SELECT COUNT(*) as c FROM drivers WHERE status='waiting' AND sort_order <= ?",
                        (row["sort_order"],)
                    ).fetchone()["c"]
                    send_position_push(name, pos)

            except Exception:
                pass
    except Exception:
        pass
    finally:
        db.close()

@app.route("/api/my_stats/<name>")
def my_stats(name):
    db = con()
    try:
        stats = compute_stats(db, name)
        return jsonify(found=True, **stats)
    except Exception as e:
        return jsonify(found=False, error=str(e))
    finally:
        db.close()

@app.route("/api/my_history/<name>")
def my_history(name):
    """آخر 5 أوردرات للسواق — بدون auth عشان السواق يقدر يشوفها من join.html."""
    db = con()
    try:
        rows = db.execute(
            "SELECT timestamp FROM done_history WHERE driver_name=? ORDER BY timestamp DESC LIMIT 5",
            (name,)
        ).fetchall()
        return jsonify(found=True, history=[r["timestamp"] for r in rows])
    except Exception as e:
        return jsonify(found=False, error=str(e))
    finally:
        db.close()

FIREBASE_WEBHOOK_SECRET = "qcd2-checkin-sync"

@app.route("/api/checkin_webhook", methods=["POST"])
def checkin_webhook():
    secret = request.headers.get("X-Webhook-Secret", "")
    if secret != FIREBASE_WEBHOOK_SECRET:
        return jsonify(error="Unauthorized"), 401
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    ticket = data.get("ticket", "")
    if not name:
        return jsonify(ok=False, error="Name empty"), 400
    db = con()
    try:
        row = db.execute("SELECT id, status FROM drivers WHERE name=?", (name,)).fetchone()
        if row and row["status"] in ("waiting", "break"):
            db.close()
            return jsonify(ok=True, skipped=True)
        max_order = db.execute("SELECT MAX(sort_order) as m FROM drivers WHERE status='waiting'").fetchone()["m"] or 0
        event_type = "checked_in"
        new_orders_count = None
        if row and row["status"] == "out":
            event_type = "new_turn"
            t = now()
            existing_order = db.execute("SELECT orders FROM driver_orders WHERE name=?", (name,)).fetchone()
            if existing_order:
                db.execute("UPDATE driver_orders SET orders=orders+1, last_done=? WHERE name=?", (t, name))
                new_orders_count = existing_order["orders"] + 1
            else:
                db.execute("INSERT INTO driver_orders(name, orders, last_done) VALUES(?,1,?)", (name, t))
                new_orders_count = 1
            db.execute("INSERT INTO done_history(driver_name, timestamp, date) VALUES(?,?,?)",
                      (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), today()))
            db.execute("UPDATE drivers SET status='waiting', created=?, out_time=NULL, done_time=?, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=? WHERE id=?",
                      (t, t, max_order + 1000, row["id"]))
        elif row:
            db.execute("UPDATE drivers SET status='waiting', created=?, out_time=NULL, done_time=NULL, break_start=NULL, break_paused_accum=0, break_pause_started=NULL, sort_order=? WHERE id=?",
                      (now(), max_order + 1000, row["id"]))
        else:
            db.execute("INSERT INTO drivers(name, status, created, sort_order) VALUES(?, 'waiting', ?, ?)",
                      (name, now(), max_order + 1000))
        db.commit()
        emit_stats_all(only_name=name if event_type != "new_turn" else None)
        socketio.emit("checkin_new", {
            "name": name,
            "ticket": ticket,
            "event_type": event_type,
            "orders_count": new_orders_count
        })
        return jsonify(ok=True, added=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

@app.route("/api/contact_phone")
def contact_phone():
    db = con()
    row = db.execute("SELECT value FROM settings WHERE key='contact_phone'").fetchone()
    db.close()
    return jsonify(phone=(row["value"] if row else ""))

@app.route("/api/contact_phone", methods=["POST"])
def set_contact_phone():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    phone = request.json.get("phone","").strip()
    db = con()
    db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('contact_phone',?)", (phone,))
    db.commit(); db.close()
    return jsonify(ok=True)

@app.route("/api/shift/status")
def shift_status():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    row = db.execute("SELECT started_at, started_at_display FROM shift_state WHERE id=1").fetchone()
    db.close()
    if row and row["started_at"]:
        return jsonify(active=True, started_at=row["started_at"], started_at_display=row["started_at_display"])
    return jsonify(active=False)

@app.route("/api/shift/start", methods=["POST"])
def shift_start():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    try:
        existing = db.execute("SELECT started_at, started_at_display FROM shift_state WHERE id=1").fetchone()
        if existing and existing["started_at"]:
            db.close()
            return jsonify(ok=True, already_active=True, started_at=existing["started_at"], started_at_display=existing["started_at_display"])
        ts_iso = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ts_display = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("INSERT OR REPLACE INTO shift_state(id, started_at, started_at_display) VALUES(1, ?, ?)", (ts_iso, ts_display))
        db.execute("DELETE FROM done_history")
        db.execute("DELETE FROM miss_history")
        db.execute("DELETE FROM return_history")
        db.commit()
        emit_stats_all()
        socketio.emit("shift_changed", {"active": True, "started_at_display": ts_display})
        return jsonify(ok=True, already_active=False, started_at=ts_iso, started_at_display=ts_display)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
@page {
size: a4 portrait;
margin: 1.5cm;
}
body {
font-family: Helvetica, Arial, sans-serif;
color: #0F172A;
background: #FFFFFF;
margin: 0;
padding: 0;
}
.header {
background-color: #0A192F;
color: #FFFFFF;
padding: 25px;
text-align: center;
margin-bottom: 20px;
}
.header h1 {
margin: 0;
font-size: 26px;
letter-spacing: 2px;
color: #FFFFFF;
}
.header p {
margin: 8px 0 0;
font-size: 12px;
color: #94A3B8;
}
.kpi-table {
width: 100%;
margin-bottom: 25px;
border-spacing: 8px;
}
.kpi-box {
background: #FFFFFF;
border: 1px solid #E2E8F0;
border-top: 4px solid #0A192F;
padding: 15px;
text-align: center;
width: 25%;
}
.kpi-box h3 {
margin: 0;
font-size: 11px;
color: #64748B;
text-transform: uppercase;
font-weight: bold;
}
.kpi-box p {
margin: 8px 0 0;
font-size: 24px;
font-weight: bold;
color: #0A192F;
}
.section-title {
font-size: 16px;
font-weight: bold;
color: #0A192F;
border-bottom: 2px solid #0A192F;
padding-bottom: 5px;
margin: 20px 0 15px 0;
}
.content-table {
width: 100%;
border-spacing: 10px;
}
.col {
width: 50%;
vertical-align: top;
}
.card {
background: #F8FAFC;
border: 1px solid #E2E8F0;
padding: 15px;
margin-bottom: 15px;
}
.data-table {
width: 100%;
border-collapse: collapse;
font-size: 11px;
margin-top: 10px;
}
.data-table th {
background: #0A192F;
color: #FFFFFF;
padding: 8px;
text-align: left;
font-size: 10px;
text-transform: uppercase;
}
.data-table td {
padding: 8px;
border-bottom: 1px solid #E2E8F0;
color: #334155;
}
.peak-table {
width: 100%;
border-collapse: collapse;
font-size: 12px;
margin-top: 10px;
}
.peak-table td {
padding: 15px;
text-align: center;
background: #FFFFFF;
border: 1px solid #E2E8F0;
width: 25%;
}
.peak-table .label {
font-size: 11px;
color: #64748B;
text-transform: uppercase;
margin-bottom: 5px;
}
.peak-table .val {
font-size: 20px;
font-weight: bold;
color: #0A192F;
}
.conclusion {
background: #F1F5F9;
padding: 20px;
border-left: 6px solid #0A192F;
font-size: 12px;
line-height: 1.6;
color: #334155;
margin-top: 20px;
}
.chart-img {
width: 100%;
max-width: 220px;
display: block;
margin: 10px auto;
}
</style>
</head>
<body>
<div class="header">
<h1>SHIFT PERFORMANCE REPORT</h1>
<p>Shift Started: {{ started_display }} | Generated: {{ now_str }}</p>
</div>
<table class="kpi-table">
<tr>
<td class="kpi-box">
<h3>Total Orders</h3>
<p>{{ summary['total'] }}</p>
</td>
<td class="kpi-box">
<h3>Delivered</h3>
<p>{{ summary['delivered'] }}</p>
</td>
<td class="kpi-box">
<h3>Success Rate</h3>
<p>{{ summary['success_rate'] }}%</p>
</td>
<td class="kpi-box">
<h3>Active Drivers</h3>
<p>{{ active_count }}</p>
</td>
</tr>
</table>
<table class="content-table">
<tr>
<td class="col">
<div class="card">
<div class="section-title">Performance by Status</div>
<img src="{{ donut_url }}" class="chart-img" />
<table class="data-table">
<tr><th>Status</th><th>Count</th><th>Rate</th></tr>
<tr><td>Delivered</td><td>{{ summary['delivered'] }}</td><td>{{ summary['success_rate'] }}%</td></tr>
<tr><td>Missed</td><td>{{ summary['missed'] }}</td><td>{{ summary['miss_rate'] }}%</td></tr>
<tr><td>Returned</td><td>{{ summary['returned'] }}</td><td>{{ summary['return_rate'] }}%</td></tr>
</table>
</div>
</td>
<td class="col">
<div class="card">
<div class="section-title">Top Drivers</div>
<img src="{{ bar_url }}" class="chart-img" />
<table class="data-table">
<tr><th>Driver</th><th>Delivered</th></tr>
{% for drv in top4 %}
<tr>
<td>{{ drv['name'] }}</td>
<td>{{ drv['delivered'] }}</td>
</tr>
{% endfor %}
</table>
</div>
</td>
</tr>
</table>
<div class="section-title">Peak Delivery Times</div>
<table class="peak-table">
<tr>
<td><div class="label">Morning</div><div class="val">{{ peak_times['morning'] }}</div></td>
<td><div class="label">Afternoon</div><div class="val">{{ peak_times['afternoon'] }}</div></td>
<td><div class="label">Evening</div><div class="val">{{ peak_times['evening'] }}</div></td>
<td><div class="label">Midnight</div><div class="val">{{ peak_times['midnight'] }}</div></td>
</tr>
</table>
<div class="conclusion">
<strong>Conclusion:</strong><br/>
During this shift (started {{ started_display }}), a total of <b>{{ summary['total'] }}</b> orders were processed.
The team achieved a success rate of <b>{{ summary['success_rate'] }}%</b>, with <b>{{ active_count }}</b> active drivers on the road.
We experienced <b>{{ summary['missed'] }}</b> missed and <b>{{ summary['returned'] }}</b> returned orders.
Focus on maintaining delivery speed and analyzing missed order patterns to optimize future performance.
</div>
</body>
</html>
"""

@app.route("/api/shift/report")
def shift_report():
    if not admin_auth(): return jsonify(error="Unauthorized"), 401
    db = con()
    ended_dt = None  # تعريف المتغير قبل الـ try
    try:
        state = db.execute("SELECT started_at, started_at_display FROM shift_state WHERE id=1").fetchone()
        if not state or not state["started_at"]:
            db.close()
            return jsonify(ok=False, error="No active shift"), 400
        started_at = state["started_at"]
        started_display = state["started_at_display"]
        ended_dt = datetime.now()  # تعريف المتغير هنا
        delivered_rows = db.execute("""
            SELECT driver_name, timestamp FROM done_history
            WHERE timestamp >= ? AND driver_name IN (SELECT email FROM allowed_accounts)
            ORDER BY timestamp ASC
        """, (started_at,)).fetchall()
        missed_rows = db.execute("""
            SELECT driver_name, order_id, reason, timestamp FROM miss_history
            WHERE timestamp >= ? AND driver_name IN (SELECT email FROM allowed_accounts)
            ORDER BY timestamp ASC
        """, (started_at,)).fetchall()
        return_rows = db.execute("""
            SELECT driver_name, order_id, timestamp FROM return_history
            WHERE timestamp >= ? AND driver_name IN (SELECT email FROM allowed_accounts)
            ORDER BY timestamp ASC
        """, (started_at,)).fetchall()
        names = set()
        for r in delivered_rows: names.add(r["driver_name"])
        for r in missed_rows: names.add(r["driver_name"])
        for r in return_rows: names.add(r["driver_name"])
        per_driver = []
        for nm in sorted(names):
            d_count = sum(1 for r in delivered_rows if r["driver_name"] == nm)
            m_count = sum(1 for r in missed_rows if r["driver_name"] == nm)
            r_count = sum(1 for r in return_rows if r["driver_name"] == nm)
            per_driver.append({"name": nm, "delivered": d_count, "missed": m_count, "returned": r_count, "total": d_count + m_count + r_count})
        per_driver.sort(key=lambda x: x["delivered"], reverse=True)
        delivered = len(delivered_rows)
        missed = len(missed_rows)
        returned = len(return_rows)
        total = delivered + missed + returned
        summary = {
            "total": total, "delivered": delivered, "missed": missed, "returned": returned,
            "success_rate": round((delivered/total*100),1) if total else 0,
            "miss_rate": round((missed/total*100),1) if total else 0,
            "return_rate": round((returned/total*100),1) if total else 0,
        }
        current_drivers = [dict(x) for x in db.execute("""
            SELECT name, status FROM drivers
            WHERE name IN (SELECT email FROM allowed_accounts)
        """)]
        active_count = sum(1 for d in current_drivers if d["status"] in ("out", "break"))
        waiting_count = sum(1 for d in current_drivers if d["status"] == "waiting")
        morning = afternoon = evening = midnight = 0
        for r in delivered_rows:
            ts = r["timestamp"]
            try:
                hour = int(ts.split(" ")[1].split(":")[0])
                if 6 <= hour < 12: morning += 1
                elif 12 <= hour < 17: afternoon += 1
                elif 17 <= hour < 21: evening += 1
                else: midnight += 1
            except:
                pass
        peak_times = {"morning": morning, "afternoon": afternoon, "evening": evening, "midnight": midnight}
        db.execute("UPDATE shift_state SET started_at=NULL, started_at_display=NULL WHERE id=1")
        db.commit()
        socketio.emit("shift_changed", {"active": False})
        pad = lambda n: str(n).zfill(2)
        fname = f"Shift_Report_{ended_dt.year}-{pad(ended_dt.month)}-{pad(ended_dt.day)}_{pad(ended_dt.hour)}-{pad(ended_dt.minute)}.xlsx"
        # ---- palette ----
        NAVY = "0A192F"
        NAVY2 = "112240"
        BLUE = "3B82F6"
        BLUE_DARK = "1E3A8A"
        GREEN = "16A34A"
        GREEN_BG = "F0FDF4"
        RED = "DC2626"
        RED_BG = "FEF2F2"
        AMBER = "D97706"
        AMBER_BG = "FFFBEB"
        SLATE = "64748B"
        LIGHT = "EFF6FF"
        WHITE = "FFFFFF"
        GREY = "F1F5F9"
        BORDER_CLR = "E2E8F0"
        TEXT = "1E293B"
        MUTED = "64748B"
        FONT_NAME = "Calibri"
        f_title = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
        f_subtitle = Font(name=FONT_NAME, size=10, color="CBD5E1")
        f_kpi_label = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        f_kpi_value = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
        f_section = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
        f_header = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        f_label = Font(name=FONT_NAME, size=11, bold=True, color=TEXT)
        f_value = Font(name=FONT_NAME, size=11, color=TEXT)
        f_normal = Font(name=FONT_NAME, size=10.5, color=TEXT)
        f_muted = Font(name=FONT_NAME, size=9.5, italic=True, color=MUTED)
        fill_navy = PatternFill("solid", fgColor=NAVY)
        fill_navy2 = PatternFill("solid", fgColor=NAVY2)
        fill_blue = PatternFill("solid", fgColor=BLUE)
        fill_blue_dark = PatternFill("solid", fgColor=BLUE_DARK)
        fill_green = PatternFill("solid", fgColor=GREEN)
        fill_red = PatternFill("solid", fgColor=RED)
        fill_amber = PatternFill("solid", fgColor=AMBER)
        fill_slate = PatternFill("solid", fgColor=SLATE)
        fill_light = PatternFill("solid", fgColor=LIGHT)
        fill_white = PatternFill("solid", fgColor=WHITE)
        fill_grey = PatternFill("solid", fgColor=GREY)
        fill_green_bg = PatternFill("solid", fgColor=GREEN_BG)
        fill_red_bg = PatternFill("solid", fgColor=RED_BG)
        fill_amber_bg = PatternFill("solid", fgColor=AMBER_BG)
        thin = Side(style="thin", color=BORDER_CLR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_bottom = Border(bottom=Side(style="medium", color=NAVY))
        center = Alignment(horizontal="center", vertical="center")
        left_a = Alignment(horizontal="left", vertical="center", indent=1)
        right_a = Alignment(horizontal="right", vertical="center", indent=1)
        def no_gridlines(ws):
            ws.sheet_view.showGridLines = False
        def style_header_row(ws, row, headers, span_widths=None):
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=i, value=h)
                c.font = f_header
                c.fill = fill_navy2
                c.alignment = center
                c.border = border
            ws.row_dimensions[row].height = 22
            ws.freeze_panes = ws.cell(row=row + 1, column=1)
        def autosize(ws, widths):
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
        def zebra(r):
            return fill_grey if r % 2 == 0 else fill_white
        def setup_print(ws):
            ws.page_setup.orientation = "landscape"
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_options.horizontalCentered = True
        wb = openpyxl.Workbook()
        # ============================================================
        # Sheet 1: Summary (dashboard style)
        # ============================================================
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = NAVY
        no_gridlines(ws)
        setup_print(ws)
        # Banner
        ws.merge_cells("A1:F1")
        c = ws.cell(row=1, column=1, value="📊  Shift Performance Report")
        c.font = f_title
        c.fill = fill_navy
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        ws.merge_cells("A2:F2")
        c = ws.cell(row=2, column=1, value=f"Shift started {started_display}   •   Report generated {ended_dt.strftime('%Y-%m-%d %H:%M')}")
        c.font = f_subtitle
        c.fill = fill_navy
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 8
        # KPI cards row (4 cards spanning columns)
        kpi_cards = [
            ("TOTAL ORDERS", summary["total"], fill_navy2, (1, 1)),
            ("DELIVERED", summary["delivered"], fill_green, (2, 2)),
            ("MISSED", summary["missed"], fill_red, (3, 3)),
            ("RETURNED", summary["returned"], fill_amber, (4, 4)),
        ]
        card_row_label, card_row_value = 4, 5
        for label, value, fill, (start_col, end_col) in kpi_cards:
            ws.merge_cells(start_row=card_row_label, start_column=start_col, end_row=card_row_label, end_column=end_col)
            lc = ws.cell(row=card_row_label, column=start_col, value=label)
            lc.font = f_kpi_label
            lc.fill = fill
            lc.alignment = center
            ws.merge_cells(start_row=card_row_value, start_column=start_col, end_row=card_row_value, end_column=end_col)
            vc = ws.cell(row=card_row_value, column=start_col, value=value)
            vc.font = f_kpi_value
            vc.fill = fill
            vc.alignment = center
        ws.row_dimensions[card_row_label].height = 20
        ws.row_dimensions[card_row_value].height = 34
        ws.row_dimensions[6].height = 10
        # Rates section
        r = 7
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value="Performance Rates")
        c.font = f_section
        c.border = thick_bottom
        r += 1
        rate_rows = [
            ("Success Rate", summary["success_rate"] / 100, GREEN, fill_green_bg),
            ("Miss Rate", summary["miss_rate"] / 100, RED, fill_red_bg),
            ("Return Rate", summary["return_rate"] / 100, AMBER, fill_amber_bg),
        ]
        for label, pct, color, bg in rate_rows:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = f_label
            lc.fill = bg
            lc.border = border
            lc.alignment = left_a
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            vc = ws.cell(row=r, column=2, value=pct)
            vc.number_format = "0.0%"
            vc.font = Font(name=FONT_NAME, size=12, bold=True, color=color)
            vc.fill = bg
            vc.border = border
            vc.alignment = center
            r += 1
        r += 1
        # Drivers + peak section side by side headers
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value="Drivers")
        c.font = f_section
        c.border = thick_bottom
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=3, value="Peak Delivery Times")
        c2.font = f_section
        c2.border = thick_bottom
        r += 1
        driver_lines = [("Active (Out / Break)", active_count), ("Waiting", waiting_count)]
        peak_lines = [("Morning (06–12)", peak_times["morning"]), ("Afternoon (12–17)", peak_times["afternoon"]),
                      ("Evening (17–21)", peak_times["evening"]), ("Midnight (21–06)", peak_times["midnight"])]
        max_lines = max(len(driver_lines), len(peak_lines))
        for i in range(max_lines):
            row_fill = zebra(r)
            if i < len(driver_lines):
                lbl, val = driver_lines[i]
                lc = ws.cell(row=r, column=1, value=lbl)
                vc = ws.cell(row=r, column=2, value=val)
                lc.font = f_value; vc.font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
                lc.fill = row_fill; vc.fill = row_fill
                lc.border = border; vc.border = border
                lc.alignment = left_a; vc.alignment = center
            if i < len(peak_lines):
                lbl, val = peak_lines[i]
                lc = ws.cell(row=r, column=3, value=lbl)
                vc = ws.cell(row=r, column=4, value=val)
                lc.font = f_value; vc.font = Font(name=FONT_NAME, size=11, bold=True, color=BLUE_DARK)
                lc.fill = row_fill; vc.fill = row_fill
                lc.border = border; vc.border = border
                lc.alignment = left_a; vc.alignment = center
            r += 1
        autosize(ws, [22, 16, 22, 16, 4, 4])
        # ============================================================
        # Sheet 2: Delivered
        # ============================================================
        ws = wb.create_sheet("Delivered")
        ws.sheet_properties.tabColor = GREEN
        no_gridlines(ws)
        setup_print(ws)
        ws.merge_cells("A1:B1")
        c = ws.cell(row=1, column=1, value=f"✅  Delivered Orders — {delivered} total")
        c.font = f_title
        c.fill = fill_green
        c.alignment = center
        ws.row_dimensions[1].height = 32
        style_header_row(ws, 2, ["Driver Name", "Time / Date"])
        r = 3
        for row in delivered_rows:
            fill = zebra(r)
            cell1 = ws.cell(row=r, column=1, value=row["driver_name"])
            cell2 = ws.cell(row=r, column=2, value=row["timestamp"])
            for cell, align in ((cell1, left_a), (cell2, center)):
                cell.font = f_normal
                cell.fill = fill
                cell.border = border
                cell.alignment = align
            r += 1
        if delivered == 0:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
            c = ws.cell(row=3, column=1, value="No delivered orders this shift")
            c.font = f_muted
            c.alignment = center
        autosize(ws, [32, 24])
        # ============================================================
        # Sheet 3: Missed
        # ============================================================
        ws = wb.create_sheet("Missed")
        ws.sheet_properties.tabColor = RED
        no_gridlines(ws)
        setup_print(ws)
        ws.merge_cells("A1:D1")
        c = ws.cell(row=1, column=1, value=f"❌  Missed Orders — {missed} total")
        c.font = f_title
        c.fill = fill_red
        c.alignment = center
        ws.row_dimensions[1].height = 32
        style_header_row(ws, 2, ["Driver Name", "Order ID", "Miss Reason", "Time / Date"])
        r = 3
        for row in missed_rows:
            vals = [row["driver_name"], row["order_id"] or "N/A", row["reason"] or "N/A", row["timestamp"]]
            fill = fill_red_bg if r % 2 == 0 else fill_white
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = f_normal
                cell.fill = fill
                cell.border = border
                cell.alignment = center if col in (2, 4) else left_a
                if col == 3:
                    cell.font = Font(name=FONT_NAME, size=10.5, color=RED, bold=True)
            r += 1
        if missed == 0:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
            c = ws.cell(row=3, column=1, value="No missed orders this shift")
            c.font = f_muted
            c.alignment = center
        autosize(ws, [30, 16, 36, 24])
        # ============================================================
        # Sheet 4: Returned
        # ============================================================
        ws = wb.create_sheet("Returned")
        ws.sheet_properties.tabColor = AMBER
        no_gridlines(ws)
        setup_print(ws)
        ws.merge_cells("A1:C1")
        c = ws.cell(row=1, column=1, value=f"↩️  Returned Orders — {returned} total")
        c.font = f_title
        c.fill = fill_amber
        c.alignment = center
        ws.row_dimensions[1].height = 32
        style_header_row(ws, 2, ["Driver Name", "Order ID", "Time / Date"])
        r = 3
        for row in return_rows:
            vals = [row["driver_name"], row["order_id"] or "N/A", row["timestamp"]]
            fill = fill_amber_bg if r % 2 == 0 else fill_white
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = f_normal
                cell.fill = fill
                cell.border = border
                cell.alignment = center if col in (2, 3) else left_a
            r += 1
        if returned == 0:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
            c = ws.cell(row=3, column=1, value="No returned orders this shift")
            c.font = f_muted
            c.alignment = center
        autosize(ws, [30, 16, 24])
        # ============================================================
        # Sheet 5: Per Driver
        # ============================================================
        ws = wb.create_sheet("Per Driver")
        ws.sheet_properties.tabColor = BLUE
        no_gridlines(ws)
        setup_print(ws)
        ws.merge_cells("A1:E1")
        c = ws.cell(row=1, column=1, value="🏆  Per-Driver Breakdown")
        c.font = f_title
        c.fill = fill_blue_dark
        c.alignment = center
        ws.row_dimensions[1].height = 32
        style_header_row(ws, 2, ["Driver Name", "Delivered", "Missed", "Returned", "Total"])
        r = 3
        medals = {3: "🥇 ", 4: "🥈 ", 5: "🥉 "}
        for d in per_driver:
            fill = zebra(r)
            name_display = medals.get(r, "") + d["name"]
            vals = [name_display, d["delivered"], d["missed"], d["returned"], d["total"]]
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.fill = fill
                cell.border = border
                cell.alignment = center if col > 1 else left_a
                if col == 1:
                    cell.font = Font(name=FONT_NAME, size=11, bold=(r in medals), color=NAVY)
                elif col == 2:
                    cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=GREEN)
                elif col == 3:
                    cell.font = Font(name=FONT_NAME, size=10.5, color=RED if d["missed"] else MUTED)
                elif col == 4:
                    cell.font = Font(name=FONT_NAME, size=10.5, color=AMBER if d["returned"] else MUTED)
                else:
                    cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=TEXT)
            r += 1
        if not per_driver:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
            c = ws.cell(row=3, column=1, value="No driver activity this shift")
            c.font = f_muted
            c.alignment = center
        autosize(ws, [32, 14, 12, 14, 12])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         headers={'Content-Disposition': f'attachment;filename={fname}'})
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

# ═══════════════════════════════════════════════
# 🔔 PWA — Manifest + Service Worker + Push
# ═══════════════════════════════════════════════

@app.after_request
def fix_headers(response):
    """يتجاوز ngrok warning ويصلح Content-Type للصور."""
    response.headers['ngrok-skip-browser-warning'] = '1'
    response.headers['Access-Control-Allow-Origin'] = '*'
    path = request.path.lower()
    if path.endswith('.png'):
        response.headers['Content-Type'] = 'image/png'
    elif path.endswith('.jpg') or path.endswith('.jpeg'):
        response.headers['Content-Type'] = 'image/jpeg'
    elif path.endswith('.svg'):
        response.headers['Content-Type'] = 'image/svg+xml'
    elif path.endswith('.webp'):
        response.headers['Content-Type'] = 'image/webp'
    return response

@app.route('/.well-known/assetlinks.json')
def asset_links_direct():
    """نرجع الـ assetlinks مباشرة كـ JSON بدون أي middleware."""
    body = json.dumps([
        {
            "relation": ["delegate_permission/common.handle_all_urls"],
            "target": {
                "namespace": "android_app",
                "package_name": "com.bankai.rwa7el",
                "sha256_cert_fingerprints": [
                    "7C:75:C5:C6:59:D7:C3:5C:8B:F6:61:9B:D6:C6:95:2B:EC:54:F5:8A:D6:54:97:EF:43:6A:A2:FE:88:1C:A5:29"
                ]
            }
        }
    ])
    return Response(body, status=200, mimetype='application/json', headers={
        'ngrok-skip-browser-warning': '1',
        'Cache-Control': 'no-cache',
        'Content-Type': 'application/json'
    })

@app.route('/manifest.json')
def serve_manifest():
    return send_file(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'manifest.json'),
        mimetype='application/manifest+json'
    )

@app.route('/api/build_id')
def api_build_id():
    """يرجع BUILD_ID فريد — الـ SW بيستخدمه لمسح الكاش القديم تلقائياً."""
    import os, hashlib
    # بيحسب hash من آخر وقت تعديل لملفات المشروع الأساسية
    files = ['app.py', 'join.html', 'admin.html', 'sw.js']
    h = hashlib.md5()
    for fname in files:
        fpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
        try:
            h.update(str(os.path.getmtime(fpath)).encode())
        except OSError:
            pass
    return jsonify(id=h.hexdigest()[:10])

@app.route('/sw.js')
def serve_sw():
    return send_file(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sw.js'),
        mimetype='application/javascript'
    )

@app.route('/api/vapid_public_key')
def vapid_public_key():
    return jsonify(key=VAPID_PUBLIC_KEY)

@app.route('/api/subscribe', methods=['POST'])
def push_subscribe():
    """الطيار بيبعت الـ push subscription بتاعته عشان نخزّنها."""
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    sub  = data.get('subscription')
    if not name or not sub:
        return jsonify(ok=False, error='missing name or subscription'), 400
    db = con()
    try:
        db.execute(
            "INSERT OR REPLACE INTO push_subscriptions(name, subscription, updated_at) VALUES(?,?,?)",
            (name, json.dumps(sub), datetime.now().isoformat())
        )
        db.commit()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

@app.route('/api/notify/<path:name>', methods=['POST'])
def push_notify(name):
    """الأدمن يبعت إشعار لطيار معيّن بالاسم."""
    if not PUSH_ENABLED:
        return jsonify(ok=False, error='push not enabled'), 503
    db = con()
    try:
        row = db.execute(
            "SELECT subscription FROM push_subscriptions WHERE name=?", (name,)
        ).fetchone()
        if not row:
            return jsonify(ok=False, error='no subscription for this driver'), 404
        sub = json.loads(row['subscription'])
        payload = json.dumps({
            "type":  "your_turn",
            "title": "🚨 دورك دلوقتي!",
            "body":  f"يا {name} — الأدمن طالبك! استعد فوراً.",
            "url":   "/join"
        }, ensure_ascii=False)
        webpush(
            subscription_info=sub,
            data=payload,
            vapid_private_key=VAPID_PRIVATE_KEY_PATH,
            vapid_claims=VAPID_CLAIMS
        )
        return jsonify(ok=True)
    except WebPushException as e:
        # لو الـ subscription انتهت (410) امسحها
        if hasattr(e, 'response') and e.response and e.response.status_code in (404, 410):
            db.execute("DELETE FROM push_subscriptions WHERE name=?", (name,))
            db.commit()
        return jsonify(ok=False, error=str(e)), 500
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500
    finally:
        db.close()

# ═══════════════════════════════════════════════
# ✏️ Text Editor — Admin Only
# ═══════════════════════════════════════════════

import re as _re

def _get_template_path(filename):
    # نستخدم نفس الطريقة اللي الـ join route بيستخدمها — بتشتغل صح
    base = app.template_folder or 'templates'
    return os.path.join(base, filename)

@app.route('/text-editor')
def text_editor_page():
    if not admin_auth():
        return "Unauthorized", 401
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'text_editor.html')
    with open(path, 'r', encoding='utf-8') as f:
        return f.read(), 200, {'Content-Type': 'text/html; charset=utf-8'}

@app.route('/api/text-editor/get', methods=['GET'])
def text_editor_get():
    if not admin_auth():
        return jsonify(error='Unauthorized'), 401
    file = request.args.get('file', 'join')  # 'join' or 'admin'
    filename = 'join.html' if file == 'join' else 'admin.html'
    path = _get_template_path(filename)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
        # Extract all unique text nodes (non-empty, non-pure-whitespace)
        # Return raw HTML for the editor to parse client-side
        return jsonify(ok=True, html=content, filename=filename)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500

@app.route('/api/text-editor/save', methods=['POST'])
def text_editor_save():
    if not admin_auth():
        return jsonify(error='Unauthorized'), 401
    data = request.get_json(silent=True) or {}
    file = data.get('file', 'join')  # 'join' or 'admin'
    replacements = data.get('replacements', [])  # [{old, new}, ...]
    filename = 'join.html' if file == 'join' else 'admin.html'
    path = _get_template_path(filename)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
        changed = 0
        import html as _html
        print(f"[TEXT-EDITOR] file={filename}, path={path}, replacements={len(replacements)}", flush=True)
        for r in replacements:
            old_txt = r.get('old', '')
            new_txt = r.get('new', '')
            if not old_txt or old_txt == new_txt:
                continue
            # جرب البحث مباشر الأول
            if old_txt in content:
                content = content.replace(old_txt, new_txt, 1)
                changed += 1
                print(f"[TEXT-EDITOR] replaced direct: {repr(old_txt[:40])}", flush=True)
            else:
                # ممكن النص في الـ HTML فيه entities — جرب escape
                escaped = _html.escape(old_txt)
                if escaped in content:
                    content = content.replace(escaped, _html.escape(new_txt), 1)
                    changed += 1
                    print(f"[TEXT-EDITOR] replaced escaped: {repr(old_txt[:40])}", flush=True)
                else:
                    print(f"[TEXT-EDITOR] NOT FOUND: {repr(old_txt[:40])}", flush=True)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
        return jsonify(ok=True, changed=changed)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=5000, debug=False)